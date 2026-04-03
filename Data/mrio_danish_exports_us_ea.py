#!/usr/bin/env python3
"""
Create an Excel file with Danish exports by ISIC sector to:
1) the United States,
2) the sum of all available Euro Area countries in the MRIO,
3) the sum of (1) and (2),
4) the US share of that sum.

Interpretation used:
- "Exports to destination" = sales from each Danish industry row to the destination's
  35 intermediate-use columns plus the destination's 5 final-demand columns.
- Euro Area membership is based on EA (20). The script automatically keeps only the
  countries that are actually present in the workbook legend.

Tested against the workbook structure in:
"ADB-MRIO-2024-August 2025.xlsx"
"""

from __future__ import annotations
import argparse
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


EA20_CODES = [
    "AUT", "BEL", "HRV", "CYP", "EST", "FIN", "FRA", "GER", "GRC", "IRE",
    "ITA", "LVA", "LTU", "LUX", "MLT", "NET", "POR", "SVK", "SVN", "SPA",
]


def excel_col_letter(col_num: int) -> str:
    result = ""
    while col_num:
        col_num, rem = divmod(col_num - 1, 26)
        result = chr(65 + rem) + result
    return result


def country_columns(country_index_zero_based: int) -> list[int]:
    """
    Return 1-based Excel column numbers for one destination economy:
    - 35 intermediate-use columns
    - 5 final-demand columns
    """
    inter_start = 5 + 35 * country_index_zero_based
    fd_start = 2630 + 5 * country_index_zero_based
    intermediate = list(range(inter_start, inter_start + 35))
    final_demand = list(range(fd_start, fd_start + 5))
    return intermediate + final_demand


def build_result_dataframe(mrio_path: Path) -> tuple[pd.DataFrame, list[str], list[str]]:
    legend = pd.read_excel(mrio_path, sheet_name="Legend", header=None, usecols="A:B")
    legend = legend.dropna(how="all")
    legend.columns = ["Code", "Country"]
    if legend.iloc[0, 0] == "Code":
        legend = legend.iloc[1:].reset_index(drop=True)

    country_codes = legend["Code"].tolist()
    country_index = {code: i for i, code in enumerate(country_codes)}

    if "DEN" not in country_index:
        raise ValueError("DEN was not found in the Legend sheet.")
    if "USA" not in country_index:
        raise ValueError("USA was not found in the Legend sheet.")

    available_ea = [code for code in EA20_CODES if code in country_index]
    missing_ea = [code for code in EA20_CODES if code not in country_index]

    den_index = country_index["DEN"]
    den_row_start = 8 + 35 * den_index   # first Danish sector row, 1-based Excel row number

    usa_cols = country_columns(country_index["USA"])
    ea_cols = [col for code in available_ea for col in country_columns(country_index[code])]

    # Read only the rows/columns we need.
    usecols = [2, 3, 4] + usa_cols + ea_cols
    usecols_excel = ",".join(excel_col_letter(c) for c in usecols)

    df = pd.read_excel(
        mrio_path,
        sheet_name="ADB MRIO 2024",
        header=None,
        skiprows=den_row_start - 1,
        nrows=35,
        usecols=usecols_excel,
    )

    sector_names = df.iloc[:, 0].astype(str)
    usa_export = df.iloc[:, 3 : 3 + len(usa_cols)].sum(axis=1)
    ea_export = df.iloc[:, 3 + len(usa_cols) :].sum(axis=1)
    total_export = usa_export + ea_export
    usa_share = np.where(total_export != 0, usa_export / total_export, np.nan)

    result = pd.DataFrame(
        {
            "ISIC sector": sector_names,
            "Danish exports to US (USD mn)": usa_export,
            "Danish exports to EA available countries (USD mn)": ea_export,
            "US + EA exports (USD mn)": total_export,
            "US share of US+EA": usa_share,
        }
    )

    return result, available_ea, missing_ea


def write_output_excel(
    result: pd.DataFrame,
    available_ea: list[str],
    missing_ea: list[str],
    output_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Exports by sector"

    # Title + notes
    ws["A1"] = "Danish exports to the US and Euro Area by ISIC sector"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = (
        "Exports are interpreted as Danish sector sales to each destination's "
        "35 intermediate-use columns plus 5 final-demand columns."
    )
    ws["A3"] = f"Available EA countries included ({len(available_ea)}): " + ", ".join(available_ea)
    ws["A4"] = "EA countries missing from workbook: " + (", ".join(missing_ea) if missing_ea else "None")

    headers = list(result.columns)
    start_row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # write data
    for i, row in enumerate(result.itertuples(index=False), start=start_row + 1):
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, float(row[1]))
        ws.cell(i, 3, float(row[2]))
        # formulas for derived values in output workbook
        ws.cell(i, 4, f"=B{i}+C{i}")
        ws.cell(i, 5, f'=IFERROR(B{i}/D{i},"")')

    # table
    last_row = start_row + len(result)
    table = Table(displayName="ExportsBySector", ref=f"A{start_row}:E{last_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # number formats
    for r in range(start_row + 1, last_row + 1):
        for c in [2, 3, 4]:
            ws.cell(r, c).number_format = '#,##0.00'
        ws.cell(r, 5).number_format = '0.0%'

    widths = {
        "A": 52,
        "B": 20,
        "C": 31,
        "D": 22,
        "E": 18,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A7"

    thin_gray = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=6, max_row=last_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 6:
                continue
            cell.border = Border(bottom=thin_gray)

    # separate metadata sheet
    meta = wb.create_sheet("Metadata")
    meta["A1"] = "Field"
    meta["B1"] = "Value"
    meta["A1"].font = meta["B1"].font = Font(bold=True)
    meta.append(["Source workbook", str(output_path.name)])
    meta.append(["EA definition requested", "EA (20)"])
    meta.append(["EA countries included", ", ".join(available_ea)])
    meta.append(["EA countries missing", ", ".join(missing_ea) if missing_ea else "None"])
    meta.append(["Units", "Millions of US$"])
    meta.append([
        "Export definition used",
        "Danish row-sector sales to destination intermediate-use columns plus destination final-demand columns",
    ])
    meta.column_dimensions["A"].width = 26
    meta.column_dimensions["B"].width = 110
    meta["B2"] = str(output_path.name)  # keep workbook-specific note simple

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "input_workbook",
        nargs="?",
        default="ADB-MRIO-2024-August 2025.xlsx",
        help="Path to the MRIO workbook",
    )
    parser.add_argument(
        "output_workbook",
        nargs="?",
        default="danish_exports_us_ea_by_isic.xlsx",
        help="Path to the output Excel workbook",
    )
    args = parser.parse_args()

    input_path = Path(args.input_workbook)
    output_path = Path(args.output_workbook)

    result, available_ea, missing_ea = build_result_dataframe(input_path)
    write_output_excel(result, available_ea, missing_ea, output_path)

    print(f"Created: {output_path}")
    print(f"Included EA countries ({len(available_ea)}): {', '.join(available_ea)}")
    if missing_ea:
        print(f"Missing EA countries: {', '.join(missing_ea)}")
    else:
        print("Missing EA countries: None")


if __name__ == "__main__":
    main()
