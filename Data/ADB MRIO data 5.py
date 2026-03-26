import os
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ============================================================
# USER INPUT
# ============================================================

DATA_FOLDER = r"\\srv9dnbfil002\userhome\kons-ana\Desktop\Data"

MRIO_FILE = os.path.join(DATA_FOLDER, "ADB-MRIO-2024-August 2025.xlsx")
TARGET_FILE = os.path.join(DATA_FOLDER, "import_export_sector_shares.xlsx")

BEA_GROSS_OUTPUT_FILE = os.path.join(DATA_FOLDER, "GrossOutput.xlsx")
BEA_VALUE_ADDED_FILE = os.path.join(DATA_FOLDER, "ValueAdded.xlsx")

YEAR = 2024

RAW_SHEET_NAME = "US raw data"
VEGT_SHEET_NAME = "US Vægt"

# ============================================================
# COUNTRY CODES
# ============================================================

EA_CODES = [
    "AUT", "BEL", "CYP", "GER", "SPA", "EST", "FIN", "FRA", "GRC", "HRV",
    "IRE", "ITA", "LTU", "LUX", "LVA", "MLT", "NET", "POR", "SVK", "SVN"
]
USA_CODE = "USA"

# ============================================================
# ORDER OF BROAD SECTORS
# ============================================================

NACE_ORDER = [
    "A Landbrug, skovbrug og fiskeri",
    "B Råstofindvinding",
    "C Fremstillingsvirksomhed (industri)",
    "D_E Forsyningsvirksomhed",
    "F Bygge- og anlægsvirksomhed",
    "G_I Handel og transport mv.",
    "J Information og kommunikation",
    "K Finansiering og forsikring",
    "L Fast ejendom",
    "M_N Erhvervsservice",
    "O_Q Offentlig administration, undervisning og sundhed",
    "R_S Kultur, fritid og anden service",
]

# ============================================================
# HELPERS
# ============================================================

def detect_mrio_sheet(xls: pd.ExcelFile) -> str:
    for s in xls.sheet_names:
        if s.lower() != "legend":
            return s
    raise ValueError("Could not find MRIO sheet.")

def find_header_row(df0: pd.DataFrame) -> tuple[int, int]:
    for i in range(min(80, df0.shape[0])):
        row = df0.iloc[i, :].astype(str).tolist()
        hits = sum(bool(re.fullmatch(r"c\d+", x.strip())) for x in row if x not in ("nan", "None"))
        if hits >= 20:
            for j, x in enumerate(row):
                if x.strip() == "c1":
                    return i, j
    raise ValueError("Could not find sector-code header row (c1,c2,...)")

def is_country_code(x) -> bool:
    if not isinstance(x, str):
        return False
    x = x.strip()
    return bool(re.fullmatch(r"[A-Z]{3}|RoW", x))

def leading_spaces(s: str) -> int:
    return len(re.match(r"^\s*", str(s)).group(0))

def replace_sheets_in_existing_workbook(target_file: str, sheet_frames: dict[str, pd.DataFrame]):
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"Target workbook not found: {target_file}")

    wb = load_workbook(target_file)

    for sheet_name in sheet_frames.keys():
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    wb.save(target_file)
    wb.close()

    with pd.ExcelWriter(target_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, df in sheet_frames.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# ============================================================
# READ ADB MRIO INTERMEDIATE BLOCK
# ============================================================

def read_intermediate_block(in_path: str, sheet_mrio: str | None = None, sheet_legend: str = "Legend"):
    xls = pd.ExcelFile(in_path, engine="openpyxl")
    if sheet_mrio is None:
        sheet_mrio = detect_mrio_sheet(xls)

    legend = pd.read_excel(in_path, sheet_name=sheet_legend, engine="openpyxl")
    legend = legend.rename(columns={legend.columns[0]: "Code", legend.columns[1]: "Country"})
    ctry_code_to_name = dict(zip(legend["Code"].astype(str), legend["Country"].astype(str)))

    df0 = pd.read_excel(in_path, sheet_name=sheet_mrio, header=None, engine="openpyxl")

    sector_code_row, col_start = find_header_row(df0)
    country_code_row = sector_code_row - 1
    sector_name_row = sector_code_row - 2
    data_start_row = sector_code_row + 1

    col_countries = df0.iloc[country_code_row, col_start:].tolist()
    col_sectors = df0.iloc[sector_code_row, col_start:].tolist()

    int_cols = []
    for k, (cc, sc) in enumerate(zip(col_countries, col_sectors)):
        if not (isinstance(sc, str) and re.fullmatch(r"c\d+", sc.strip())):
            break
        if not is_country_code(str(cc)):
            break
        int_cols.append(col_start + k)

    row_sectorname_col = col_start - 3
    row_country_col = col_start - 2
    row_sectorcode_col = col_start - 1

    sector_codes_35 = [str(x).strip() for x in df0.iloc[sector_code_row, col_start:col_start+35].tolist()]
    sector_names_35 = [str(x).strip() for x in df0.iloc[sector_name_row, col_start:col_start+35].tolist()]
    sector_lookup = dict(zip(sector_codes_35, sector_names_35))

    rows = []
    row_keys = []
    for i in range(data_start_row, df0.shape[0]):
        cc = df0.iat[i, row_country_col]
        sc = df0.iat[i, row_sectorcode_col]
        sn = df0.iat[i, row_sectorname_col]

        if not is_country_code(str(cc)):
            break
        if not (isinstance(sc, str) and re.fullmatch(r"c\d+", sc.strip())):
            break
        if pd.isna(sn):
            break

        rows.append(i)
        row_keys.append((str(cc).strip(), str(sc).strip()))

    Z = df0.iloc[rows, int_cols].copy()
    Z = Z.apply(pd.to_numeric, errors="coerce")

    Z.index = pd.MultiIndex.from_tuples(row_keys, names=["exp_country", "exp_sector"])
    col_keys = [(str(df0.iat[country_code_row, j]).strip(), str(df0.iat[sector_code_row, j]).strip()) for j in int_cols]
    Z.columns = pd.MultiIndex.from_tuples(col_keys, names=["imp_country", "imp_sector"])

    return Z, sector_lookup, ctry_code_to_name

def to_long(Z: pd.DataFrame, sector_lookup: dict, ctry_code_to_name: dict,
            exporters: list[str], importers: list[str], label: str) -> pd.DataFrame:
    sub = Z.loc[
        Z.index.get_level_values("exp_country").isin(exporters),
        Z.columns.get_level_values("imp_country").isin(importers)
    ]

    long = sub.stack(["imp_country", "imp_sector"]).reset_index()
    long = long.rename(columns={0: "value"})
    long["scenario"] = label
    long["exp_country_name"] = long["exp_country"].map(ctry_code_to_name)
    long["imp_country_name"] = long["imp_country"].map(ctry_code_to_name)
    long["exp_sector_name"] = long["exp_sector"].map(sector_lookup)
    long["imp_sector_name"] = long["imp_sector"].map(sector_lookup)
    return long

# ============================================================
# READ MAPPING FROM TARGET WORKBOOK
# ============================================================

def read_isic_to_nace_mapping(target_file: str) -> pd.DataFrame:
    df = pd.read_excel(target_file, sheet_name="Import tabel", engine="openpyxl")
    out = df[["ISIC_sector_code", "ISIC_sector", "NACE_sector_name"]].copy()
    out = out.dropna(subset=["ISIC_sector_code", "NACE_sector_name"])
    out["ISIC_sector_code"] = out["ISIC_sector_code"].astype(str).str.strip()
    out["ISIC_sector"] = out["ISIC_sector"].astype(str).str.strip()
    out["NACE_sector_name"] = out["NACE_sector_name"].astype(str).str.strip()
    out = out.drop_duplicates()
    return out

# ============================================================
# BUILD U.S. INPUT DATA FROM MRIO
# ============================================================

def build_us_inputs_from_mrio(mrio_file: str, target_file: str):
    Z, sector_lookup, ctry_code_to_name = read_intermediate_block(mrio_file)
    map_df = read_isic_to_nace_mapping(target_file)

    long_us_inputs = to_long(
        Z=Z,
        sector_lookup=sector_lookup,
        ctry_code_to_name=ctry_code_to_name,
        exporters=EA_CODES + [USA_CODE],
        importers=[USA_CODE],
        label="US_inputs_from_EA_and_US"
    )

    long_us_inputs["origin_group"] = np.where(
        long_us_inputs["exp_country"] == USA_CODE,
        "US",
        "EA"
    )

    # raw rows by source country and U.S. sector
    raw_country = (
        long_us_inputs
        .groupby(
            ["exp_country", "exp_country_name", "origin_group", "imp_sector", "imp_sector_name"],
            as_index=False
        )["value"]
        .sum(min_count=1)
    )

    # EA sum
    ea_sum = (
        raw_country[raw_country["origin_group"] == "EA"]
        .groupby(["origin_group", "imp_sector", "imp_sector_name"], as_index=False)["value"]
        .sum(min_count=1)
    )
    ea_sum["exp_country"] = "EA"
    ea_sum["exp_country_name"] = "Euro Area"

    # U.S. self-inputs
    us_sum = (
        raw_country[raw_country["origin_group"] == "US"]
        .groupby(["origin_group", "imp_sector", "imp_sector_name"], as_index=False)["value"]
        .sum(min_count=1)
    )
    us_sum["exp_country"] = "USA"
    us_sum["exp_country_name"] = "United States"

    raw_inputs = pd.concat([raw_country, ea_sum, us_sum], ignore_index=True, sort=False)

    raw_inputs = raw_inputs.merge(
        map_df,
        left_on="imp_sector",
        right_on="ISIC_sector_code",
        how="left"
    )

    raw_inputs["source_block"] = "ADB MRIO"
    raw_inputs["metric"] = "Intermediate input to U.S."
    raw_inputs["year"] = YEAR

    raw_inputs = raw_inputs[
        [
            "source_block",
            "exp_country",
            "exp_country_name",
            "origin_group",
            "imp_sector",
            "imp_sector_name",
            "NACE_sector_name",
            "metric",
            "year",
            "value"
        ]
    ].rename(columns={
        "exp_country": "source_country_code",
        "exp_country_name": "source_country",
        "imp_sector": "ISIC_sector_code",
        "imp_sector_name": "ISIC_sector",
        "value": "value_usd"
    })

    broad = (
        raw_inputs[raw_inputs["origin_group"].isin(["EA", "US"])]
        .groupby(["NACE_sector_name", "origin_group"], as_index=False)["value_usd"]
        .sum(min_count=1)
        .pivot(index="NACE_sector_name", columns="origin_group", values="value_usd")
        .reset_index()
        .rename_axis(None, axis=1)
    )

    if "EA" not in broad.columns:
        broad["EA"] = np.nan
    if "US" not in broad.columns:
        broad["US"] = np.nan

    broad = broad.rename(columns={
        "EA": "input_EA",
        "US": "input_US"
    })

    return raw_inputs, broad

# ============================================================
# PARSE BEA ANNUAL TABLES
# ============================================================

def detect_year_col(df: pd.DataFrame, year: int) -> int:
    header = df.iloc[7].tolist()
    target = str(year)
    for i, v in enumerate(header):
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
        if s == target:
            return i
    raise ValueError(f"Could not find year column {year} in BEA sheet.")

def parse_bea_gross_output(path: str, year: int) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="TGO105-A", header=None, engine="openpyxl")
    year_col = detect_year_col(df, year)

    rows = []
    for idx in range(8, len(df)):
        desc = df.iat[idx, 1]
        if pd.isna(desc):
            continue

        desc = str(desc)
        rows.append({
            "industry_raw": desc.strip(),
            "level": leading_spaces(desc),
            "Production USD": pd.to_numeric(df.iat[idx, year_col], errors="coerce")
        })

    out = pd.DataFrame(rows)

    # leaf node if next row is same or lower indentation
    levels = out["level"].tolist()
    leaf = []
    for i, lvl in enumerate(levels):
        if i == len(levels) - 1:
            leaf.append(True)
        else:
            leaf.append(levels[i + 1] <= lvl)
    out["leaf"] = leaf

    return out

def parse_bea_compensation(path: str, year: int) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="TVA113-A", header=None, engine="openpyxl")
    year_col = detect_year_col(df, year)

    component_rows = {
        "Compensation of employees",
        "Taxes on production and imports less subsidies",
        "Gross operating surplus"
    }

    industry_rows = []
    comp_rows = []

    current_industry = None
    current_level = None

    for idx in range(8, len(df)):
        desc = df.iat[idx, 1]
        if pd.isna(desc):
            continue

        desc = str(desc)
        desc_clean = desc.strip()
        value = pd.to_numeric(df.iat[idx, year_col], errors="coerce")
        level = leading_spaces(desc)

        if desc_clean in component_rows:
            comp_rows.append({
                "industry_raw": current_industry,
                "level": current_level,
                "component": desc_clean,
                "value": value
            })
        else:
            current_industry = desc_clean
            current_level = level
            industry_rows.append({
                "industry_raw": current_industry,
                "level": current_level
            })

    ind = pd.DataFrame(industry_rows).drop_duplicates().reset_index(drop=True)

    levels = ind["level"].tolist()
    leaf = []
    for i, lvl in enumerate(levels):
        if i == len(levels) - 1:
            leaf.append(True)
        else:
            leaf.append(levels[i + 1] <= lvl)
    ind["leaf"] = leaf

    comp = pd.DataFrame(comp_rows)
    comp = comp.merge(ind, on=["industry_raw", "level"], how="left")
    comp = comp[comp["component"] == "Compensation of employees"].copy()
    comp = comp.rename(columns={"value": "Labour Costs USD"})

    return comp

# ============================================================
# MAP BEA INDUSTRIES TO BROAD SECTORS
# ============================================================

def map_bea_to_nace(industry_raw: str):
    s = str(industry_raw).strip().lower()

    # A
    if "farm" in s or "forestry" in s or "fishing" in s:
        return "A Landbrug, skovbrug og fiskeri"

    # B
    if s in {"mining", "oil and gas extraction", "mining, except oil and gas", "support activities for mining"}:
        return "B Råstofindvinding"

    # D_E
    if s == "utilities":
        return "D_E Forsyningsvirksomhed"

    # F
    if s == "construction":
        return "F Bygge- og anlægsvirksomhed"

    # C
    manufacturing_terms = [
        "wood products",
        "nonmetallic mineral products",
        "primary metals",
        "fabricated metal products",
        "machinery",
        "computer and electronic products",
        "electrical equipment, appliances, and components",
        "motor vehicles, bodies and trailers, and parts",
        "other transportation equipment",
        "furniture and related products",
        "miscellaneous manufacturing",
        "food and beverage and tobacco products",
        "textile mills and textile product mills",
        "apparel and leather and allied products",
        "paper products",
        "printing and related support activities",
        "petroleum and coal products",
        "chemical products",
        "plastics and rubber products",
    ]
    if s == "manufacturing" or s in manufacturing_terms:
        return "C Fremstillingsvirksomhed (industri)"

    # G_I
    if s in {
        "wholesale trade",
        "retail trade",
        "air transportation",
        "rail transportation",
        "water transportation",
        "truck transportation",
        "transit and ground passenger transportation",
        "pipeline transportation",
        "other transportation and support activities",
        "warehousing and storage",
        "accommodation",
        "food services and drinking places",
    }:
        return "G_I Handel og transport mv."

    # J
    if s in {
        "publishing industries, except internet (includes software)",
        "motion picture and sound recording industries",
        "broadcasting and telecommunications",
        "data processing, internet publishing, and other information services",
        "information",
    }:
        return "J Information og kommunikation"

    # K
    if s in {
        "federal reserve banks, credit intermediation, and related activities",
        "securities, commodity contracts, and investments",
        "insurance carriers and related activities",
        "funds, trusts, and other financial vehicles",
        "finance and insurance",
    }:
        return "K Finansiering og forsikring"

    # L
    if s in {
        "real estate",
        "rental and leasing services and lessors of intangible assets",
        "other real estate",
        "owner-occupied housing",
        "real estate and rental and leasing",
    }:
        return "L Fast ejendom"

    # M_N
    if s in {
        "legal services",
        "computer systems design and related services",
        "miscellaneous professional, scientific, and technical services",
        "management of companies and enterprises",
        "administrative and support services",
        "waste management and remediation services",
        "professional, scientific, and technical services",
        "administrative and waste management services",
    }:
        return "M_N Erhvervsservice"

    # O_Q
    if s in {
        "educational services",
        "ambulatory health care services",
        "hospitals and nursing and residential care facilities",
        "social assistance",
        "health care and social assistance",
        "general government",
        "government enterprises",
        "government",
    }:
        return "O_Q Offentlig administration, undervisning og sundhed"

    # R_S
    if s in {
        "performing arts, spectator sports, museums, and related activities",
        "amusements, gambling, and recreation industries",
        "arts, entertainment, and recreation",
        "repair and maintenance",
        "personal and laundry services",
        "religious, grantmaking, civic, professional, and similar organizations",
        "other services, except government",
    }:
        return "R_S Kultur, fritid og anden service"

    return None

# ============================================================
# BUILD BEA U.S. PRODUCTION + LABOUR DATA
# ============================================================

def build_bea_us_data(gross_output_file: str, value_added_file: str):
    gross = parse_bea_gross_output(gross_output_file, YEAR)
    comp = parse_bea_compensation(value_added_file, YEAR)

    gross = gross[gross["leaf"]].copy()
    comp = comp[comp["leaf"]].copy()

    gross["NACE_sector_name"] = gross["industry_raw"].map(map_bea_to_nace)
    comp["NACE_sector_name"] = comp["industry_raw"].map(map_bea_to_nace)

    gross = gross.dropna(subset=["NACE_sector_name"]).copy()
    comp = comp.dropna(subset=["NACE_sector_name"]).copy()

    raw_gross = gross[["industry_raw", "NACE_sector_name", "Production USD"]].copy()
    raw_gross["source_block"] = "BEA GrossOutput TGO105-A"
    raw_gross["metric"] = "Production USD"
    raw_gross["year"] = YEAR
    raw_gross = raw_gross.rename(columns={"Production USD": "value_usd"})

    raw_comp = comp[["industry_raw", "NACE_sector_name", "Labour Costs USD"]].copy()
    raw_comp["source_block"] = "BEA ValueAdded TVA113-A"
    raw_comp["metric"] = "Labour Costs USD"
    raw_comp["year"] = YEAR
    raw_comp = raw_comp.rename(columns={"Labour Costs USD": "value_usd"})

    raw_bea = pd.concat([raw_gross, raw_comp], ignore_index=True, sort=False)

    prod = (
        gross.groupby("NACE_sector_name", as_index=False)["Production USD"]
        .sum(min_count=1)
    )
    lab = (
        comp.groupby("NACE_sector_name", as_index=False)["Labour Costs USD"]
        .sum(min_count=1)
    )

    bea_agg = prod.merge(lab, on="NACE_sector_name", how="outer")
    return raw_bea, bea_agg

# ============================================================
# MAIN
# ============================================================

def main():
    # 1. Inputs into U.S. sectors from EA and U.S.
    raw_inputs, broad_inputs = build_us_inputs_from_mrio(MRIO_FILE, TARGET_FILE)

    # 2. U.S. production and labour costs from BEA
    raw_bea, bea_agg = build_bea_us_data(BEA_GROSS_OUTPUT_FILE, BEA_VALUE_ADDED_FILE)

    # 3. Build U.S. summary sheet
    us_vaegt = broad_inputs.merge(bea_agg, on="NACE_sector_name", how="outer")

    # sort rows
    order_map = {name: i for i, name in enumerate(NACE_ORDER, start=1)}
    us_vaegt["sort_order"] = us_vaegt["NACE_sector_name"].map(order_map).fillna(999)
    us_vaegt = us_vaegt.sort_values(["sort_order", "NACE_sector_name"]).drop(columns="sort_order")

    # 4. Raw sheet
    raw_inputs = raw_inputs.copy()
    raw_bea = raw_bea.copy()

    raw_inputs["industry_raw"] = np.nan
    raw_inputs = raw_inputs[
        [
            "source_block",
            "source_country_code",
            "source_country",
            "origin_group",
            "ISIC_sector_code",
            "ISIC_sector",
            "NACE_sector_name",
            "metric",
            "year",
            "industry_raw",
            "value_usd"
        ]
    ]

    raw_bea["source_country_code"] = "USA"
    raw_bea["source_country"] = "United States"
    raw_bea["origin_group"] = "US macro data"
    raw_bea["ISIC_sector_code"] = np.nan
    raw_bea["ISIC_sector"] = np.nan

    raw_bea = raw_bea[
        [
            "source_block",
            "source_country_code",
            "source_country",
            "origin_group",
            "ISIC_sector_code",
            "ISIC_sector",
            "NACE_sector_name",
            "metric",
            "year",
            "industry_raw",
            "value_usd"
        ]
    ]

    us_raw = pd.concat([raw_inputs, raw_bea], ignore_index=True, sort=False)

    # 5. Save
    replace_sheets_in_existing_workbook(
        TARGET_FILE,
        {
            RAW_SHEET_NAME: us_raw,
            VEGT_SHEET_NAME: us_vaegt
        }
    )

    print("Updated workbook:", TARGET_FILE)
    print("Added/replaced sheets:")
    print(" -", RAW_SHEET_NAME)
    print(" -", VEGT_SHEET_NAME)

if __name__ == "__main__":
    main()