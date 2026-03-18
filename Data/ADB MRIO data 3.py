import re
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# =========================================================
# USER INPUT
# =========================================================

IN_PATH = r"\\srv9dnbfil002\userhome\kons-ana\Desktop\Data\ADB-MRIO-2024-August 2025.xlsx"
SHEET_MRIO = None
SHEET_LEGEND = "Legend"

# Existing workbook to amend
out_folder = os.path.dirname(IN_PATH)
TARGET_FILE = os.path.join(out_folder, "adb_mrio_intermediate_imports_long.xlsx")

# Optional DST employment file (10a3)
DST_EMPLOYMENT_FILE = r"\\srv9dnbfil002\userhome\kons-ana\Desktop\Data\202631313395608068025NABB10.xlsx"

# =========================================================
# CONSTANTS
# =========================================================

EA20 = {
    "Austria","Belgium","Croatia","Cyprus","Estonia","Finland","France","Germany",
    "Greece","Ireland","Italy","Latvia","Lithuania","Luxembourg","Malta","Netherlands",
    "Portugal","Slovakia","Slovenia","Spain"
}

EA_CODE = "EA"
EA_NAME = "Euro area (EA20)"

USA_NAME = "United States"
DEN_NAME = "Denmark"

NEW_SHEETS = [
    "EXP_USA_ISIC_by_exporter",
    "EXP_USA_NACE_by_exporter",
    "EXP_USA_10a3_by_exporter",
    "EXP_USA_exporter_totals",
    "EXP_USA_10a3_exposure"
]

# =========================================================
# HELPERS
# =========================================================

def clean_str(x):
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    s = re.sub(r"\s+", " ", s)
    return s

def detect_mrio_sheet(xls: pd.ExcelFile) -> str:
    for s in xls.sheet_names:
        if s.lower() != "legend":
            return s
    raise ValueError("Could not find the MRIO sheet (other than 'Legend').")

def find_header_row(df0: pd.DataFrame):
    for i in range(min(80, df0.shape[0])):
        row = df0.iloc[i, :].astype(str).tolist()
        hits = sum(bool(re.fullmatch(r"c\d+", x.strip())) for x in row if x not in ("nan", "None"))
        if hits >= 20:
            for j, x in enumerate(row):
                if x.strip() == "c1":
                    return i, j
    raise ValueError("Could not find sector-code header row (c1,c2,...) in the first 80 rows.")

def is_country_code(x) -> bool:
    if not isinstance(x, str):
        return False
    x = x.strip()
    return bool(re.fullmatch(r"[A-Z]{3}|RoW", x))

def replace_sheets_in_existing_workbook(target_file: str, sheet_frames: dict):
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"Target workbook not found: {target_file}")

    wb = load_workbook(target_file)

    for s in sheet_frames.keys():
        if s in wb.sheetnames:
            del wb[s]

    wb.save(target_file)
    wb.close()

    with pd.ExcelWriter(
        target_file,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        for sheet_name, df in sheet_frames.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# =========================================================
# READ ADB MRIO
# =========================================================

def read_mrio_blocks(in_path: str, sheet_mrio: str, sheet_legend: str):
    xls = pd.ExcelFile(in_path, engine="openpyxl")
    if sheet_mrio is None:
        sheet_mrio = detect_mrio_sheet(xls)

    legend = pd.read_excel(in_path, sheet_name=sheet_legend, engine="openpyxl")
    legend = legend.rename(columns={legend.columns[0]: "Code", legend.columns[1]: "Country"})
    legend["Code"] = legend["Code"].astype(str).str.strip()
    legend["Country"] = legend["Country"].astype(str).str.strip()

    ctry_code_to_name = dict(zip(legend["Code"], legend["Country"]))
    ctry_name_to_code = dict(zip(legend["Country"], legend["Code"]))

    df0 = pd.read_excel(in_path, sheet_name=sheet_mrio, header=None, engine="openpyxl")

    sector_code_row, col_start = find_header_row(df0)
    country_code_row = sector_code_row - 1
    sector_name_row = sector_code_row - 2
    data_start_row = sector_code_row + 1

    row_sectorname_col = col_start - 3
    row_country_col = col_start - 2
    row_sectorcode_col = col_start - 1

    # Intermediate columns
    int_cols = []
    for k in range(col_start, df0.shape[1]):
        cc = df0.iat[country_code_row, k]
        sc = df0.iat[sector_code_row, k]

        if not (isinstance(sc, str) and re.fullmatch(r"c\d+", sc.strip())):
            break
        if not is_country_code(str(cc)):
            break
        int_cols.append(k)

    if len(int_cols) < 200:
        raise ValueError(f"Found only {len(int_cols)} intermediate columns — layout may differ.")

    # Final demand columns
    fd_cols = []
    k = int_cols[-1] + 1
    while k < df0.shape[1]:
        cc = df0.iat[country_code_row, k]
        fc = df0.iat[sector_code_row, k]

        if not (isinstance(fc, str) and re.fullmatch(r"F[1-5]", fc.strip())):
            break
        if not is_country_code(str(cc)):
            break
        fd_cols.append(k)
        k += 1

    if len(fd_cols) == 0:
        raise ValueError("Found no final-demand columns (F1-F5).")

    sector_codes_35 = [str(x).strip() for x in df0.iloc[sector_code_row, col_start:col_start+35].tolist()]
    sector_names_35 = [str(x).strip() for x in df0.iloc[sector_name_row, col_start:col_start+35].tolist()]
    sector_lookup = dict(zip(sector_codes_35, sector_names_35))

    rows = []
    row_keys = []

    for i in range(data_start_row, df0.shape[0]):
        cc = df0.iat[i, row_country_col]
        sc = df0.iat[i, row_sectorcode_col]
        sn = df0.iat[i, row_sectorname_col]

        if not is_country_code(str(cc)):
            break
        if not (isinstance(sc, str) and re.fullmatch(r"c\d+", sc.strip())):
            break
        if pd.isna(sn):
            break

        rows.append(i)
        row_keys.append((str(cc).strip(), str(sc).strip()))

    if len(rows) < 200:
        raise ValueError(f"Found only {len(rows)} sector rows — layout may differ.")

    Z = df0.iloc[rows, int_cols].copy()
    Z = Z.apply(pd.to_numeric, errors="coerce")
    Z.index = pd.MultiIndex.from_tuples(row_keys, names=["exp_country", "exp_sector"])
    Z.columns = pd.MultiIndex.from_tuples(
        [(str(df0.iat[country_code_row, j]).strip(), str(df0.iat[sector_code_row, j]).strip()) for j in int_cols],
        names=["imp_country", "imp_sector"]
    )

    F = df0.iloc[rows, fd_cols].copy()
    F = F.apply(pd.to_numeric, errors="coerce")
    F.index = pd.MultiIndex.from_tuples(row_keys, names=["exp_country", "exp_sector"])
    F.columns = pd.MultiIndex.from_tuples(
        [(str(df0.iat[country_code_row, j]).strip(), str(df0.iat[sector_code_row, j]).strip()) for j in fd_cols],
        names=["fd_country", "fd_use"]
    )

    return Z, F, sector_lookup, ctry_code_to_name, ctry_name_to_code

# =========================================================
# EXTRACT EXPORTS TO USA:
#   DK -> USA
#   each EA country -> USA
#   EA total -> USA
# =========================================================

def extract_exports_to_usa_by_exporter(Z, F, sector_lookup, ctry_code_to_name, ctry_name_to_code):
    DEN = ctry_name_to_code.get(DEN_NAME, "DEN")
    USA = ctry_name_to_code.get(USA_NAME, "USA")

    present_countries = set(ctry_name_to_code.keys())
    ea_present_names = sorted(EA20.intersection(present_countries))
    ea_present_codes = [ctry_name_to_code[n] for n in ea_present_names]

    exporter_codes = [DEN] + ea_present_codes
    exporter_names = {DEN: DEN_NAME}
    for n in ea_present_names:
        exporter_names[ctry_name_to_code[n]] = n

    out_rows = []

    # destination is fixed = USA
    for exporter_code in exporter_codes:
        z_rows = Z.index.get_level_values("exp_country") == exporter_code
        Z_exp = Z.loc[z_rows].copy()
        F_exp = F.loc[z_rows].copy()

        # exports to USA intermediate uses
        z_sub = Z_exp.loc[:, Z_exp.columns.get_level_values("imp_country") == USA]
        if z_sub.shape[1] > 0:
            z_sum = z_sub.groupby(level="exp_sector").sum(min_count=1).sum(axis=1)
        else:
            z_sum = pd.Series(dtype=float)

        # exports to USA final demand
        f_sub = F_exp.loc[:, F_exp.columns.get_level_values("fd_country") == USA]
        if f_sub.shape[1] > 0:
            f_sum = f_sub.groupby(level="exp_sector").sum(min_count=1).sum(axis=1)
        else:
            f_sum = pd.Series(dtype=float)

        idx = sorted(set(z_sum.index).union(set(f_sum.index)))

        for sec in idx:
            intermediate_val = float(z_sum.get(sec, 0.0)) if pd.notna(z_sum.get(sec, 0.0)) else 0.0
            final_val = float(f_sum.get(sec, 0.0)) if pd.notna(f_sum.get(sec, 0.0)) else 0.0

            out_rows.append({
                "exp_country_code": exporter_code,
                "exp_country_name": exporter_names.get(exporter_code, ctry_code_to_name.get(exporter_code, exporter_code)),
                "dest_country_code": USA,
                "dest_country_name": USA_NAME,
                "ISIC_sector_code": sec,
                "ISIC_sector": sector_lookup.get(sec, sec),
                "exports_intermediate": intermediate_val,
                "exports_final": final_val,
                "exports_total": intermediate_val + final_val
            })

    isic_df = pd.DataFrame(out_rows)

    # EA total -> USA
    ea_df = isic_df[isic_df["exp_country_code"].isin(ea_present_codes)].copy()
    ea_total = (
        ea_df.groupby(["dest_country_code", "dest_country_name", "ISIC_sector_code", "ISIC_sector"], as_index=False)
        [["exports_intermediate", "exports_final", "exports_total"]]
        .sum(min_count=1)
    )
    ea_total["exp_country_code"] = EA_CODE
    ea_total["exp_country_name"] = EA_NAME

    isic_df = pd.concat([isic_df, ea_total], ignore_index=True)

    isic_df = isic_df[
        [
            "exp_country_code", "exp_country_name",
            "dest_country_code", "dest_country_name",
            "ISIC_sector_code", "ISIC_sector",
            "exports_intermediate", "exports_final", "exports_total"
        ]
    ].sort_values(["exp_country_name", "ISIC_sector_code"]).reset_index(drop=True)

    return isic_df, ea_present_names, ea_present_codes

# =========================================================
# LOAD EXISTING WORKBOOK DATA
# =========================================================

def load_existing_mapping_and_nace(target_file):
    mapping = pd.read_excel(target_file, sheet_name="Mapping")
    nace_table = pd.read_excel(target_file, sheet_name="NACE tabel")

    mapping.columns = [clean_str(c) for c in mapping.columns]
    nace_table.columns = [clean_str(c) for c in nace_table.columns]

    required_mapping_cols = [
        "ISIC_sector_code",
        "ISIC_sector",
        "NACE_sector_code",
        "NACE_sector_name"
    ]
    for c in required_mapping_cols:
        if c not in mapping.columns:
            raise ValueError(f"Missing column '{c}' in sheet 'Mapping'.")

    required_nace_cols = ["NACE_sector_name", "Production", "NVA"]
    for c in required_nace_cols:
        if c not in nace_table.columns:
            raise ValueError(f"Missing column '{c}' in sheet 'NACE tabel'.")

    mapping_small = mapping[required_mapping_cols].copy()
    mapping_small["ISIC_sector_code"] = mapping_small["ISIC_sector_code"].astype(str).str.strip()
    mapping_small["ISIC_sector"] = mapping_small["ISIC_sector"].map(clean_str)
    mapping_small["NACE_sector_code"] = mapping_small["NACE_sector_code"].map(clean_str)
    mapping_small["NACE_sector_name"] = mapping_small["NACE_sector_name"].map(clean_str)

    nace_small = nace_table.copy()
    nace_small["NACE_sector_name"] = nace_small["NACE_sector_name"].map(clean_str)

    return mapping_small, nace_small

# =========================================================
# MAP ISIC -> NACE
# =========================================================

def build_nace_exports(isic_exports, mapping_small):
    tmp = mapping_small.merge(
        isic_exports,
        on=["ISIC_sector_code", "ISIC_sector"],
        how="left"
    )

    for c in ["exports_intermediate", "exports_final", "exports_total"]:
        tmp[c] = pd.to_numeric(tmp[c], errors="coerce").fillna(0.0)

    nace_exports = (
        tmp.groupby(["exp_country_code", "exp_country_name", "dest_country_code", "dest_country_name",
                     "NACE_sector_code", "NACE_sector_name"], as_index=False)
        [["exports_intermediate", "exports_final", "exports_total"]]
        .sum(min_count=1)
    )

    nace_exports = nace_exports.sort_values(["exp_country_name", "NACE_sector_code"]).reset_index(drop=True)
    return nace_exports

# =========================================================
# MAP NACE -> 10A3
# =========================================================

def build_nace_to_10a3_map():
    return {
        "A Landbrug, skovbrug og fiskeri": ["A Landbrug, skovbrug og fiskeri"],
        "B Råstofindvinding": ["B Råstofindvinding"],
        "C Fremstillingsvirksomhed (industri)": ["C Industri"],
        "D_E Forsyningsvirksomhed": ["D_E Forsyningsvirksomhed"],
        "F Bygge- og anlægsvirksomhed": ["F Bygge og anlæg"],
        "G_I Handel og transport mv.": ["G_I Handel og transport mv."],
        "J Information og kommunikation": ["J Information og kommunikation"],
        "K Finansiering og forsikring": ["K Finansiering og forsikring"],
        "L Fast ejendom": ["LA Ejendomshandel og udlejning af erhvervsejendomme", "LB Boliger"],
        "M_N Erhvervsservice": ["M_N Erhvervsservice"],
        "O_Q Offentlig administration, undervisning og sundhed": ["O_Q Offentlig administration, undervisning og sundhed"],
        "R_S Kultur, fritid og anden service": ["R_S Kultur, fritid og anden service"],
    }

def load_dst_10a3_employment(dst_file):
    if not os.path.exists(dst_file):
        return None

    raw = pd.read_excel(dst_file, sheet_name=0, header=None)

    data = raw.iloc[3:].copy()
    data.columns = ["year", "branch_10a3", "employment_total", "employment_wage_earners"]

    data = data[["branch_10a3", "employment_total", "employment_wage_earners"]].copy()
    data["branch_10a3"] = data["branch_10a3"].map(clean_str)
    data["employment_total"] = pd.to_numeric(data["employment_total"], errors="coerce")
    data["employment_wage_earners"] = pd.to_numeric(data["employment_wage_earners"], errors="coerce")

    data = data[~data["branch_10a3"].isna()].copy()
    data = data[data["branch_10a3"] != "Heraf: Offentlig forvaltning og service"].copy()

    return data

def build_10a3_exports_and_exposure(nace_exports, nace_small, dst_emp=None):
    nace_to_10a3 = build_nace_to_10a3_map()

    rows = []
    for _, row in nace_exports.iterrows():
        nace_name = row["NACE_sector_name"]
        if nace_name not in nace_to_10a3:
            continue

        target_groups = nace_to_10a3[nace_name]

        if len(target_groups) == 1:
            out = row.to_dict()
            out["branch_10a3"] = target_groups[0]
            out["allocation_share"] = 1.0
            rows.append(out)
        else:
            if dst_emp is not None:
                subset = dst_emp[dst_emp["branch_10a3"].isin(target_groups)].copy()
                subset = subset.dropna(subset=["employment_total"])
            else:
                subset = pd.DataFrame()

            if subset.empty or subset["employment_total"].sum() == 0:
                shares = {target_groups[0]: 0.5, target_groups[1]: 0.5}
            else:
                shares = (
                    subset.set_index("branch_10a3")["employment_total"] /
                    subset["employment_total"].sum()
                ).to_dict()

            for g in target_groups:
                out = row.to_dict()
                out["branch_10a3"] = g
                out["allocation_share"] = shares.get(g, 0.0)
                rows.append(out)

    exp_10a3 = pd.DataFrame(rows)

    for c in ["exports_intermediate", "exports_final", "exports_total"]:
        exp_10a3[c] = exp_10a3[c] * exp_10a3["allocation_share"]

    exp_10a3 = (
        exp_10a3.groupby(["exp_country_code", "exp_country_name", "dest_country_code", "dest_country_name", "branch_10a3"], as_index=False)
        [["exports_intermediate", "exports_final", "exports_total"]]
        .sum(min_count=1)
    )

    exposure = None
    if dst_emp is not None:
        # Build base Production/NVA at 10a3
        base_rows = []
        nace_base = nace_small[["NACE_sector_name", "Production", "NVA"]].copy()

        for _, row in nace_base.iterrows():
            nace_name = row["NACE_sector_name"]
            if nace_name not in nace_to_10a3:
                continue

            target_groups = nace_to_10a3[nace_name]

            if len(target_groups) == 1:
                out = row.to_dict()
                out["branch_10a3"] = target_groups[0]
                out["allocation_share"] = 1.0
                base_rows.append(out)
            else:
                subset = dst_emp[dst_emp["branch_10a3"].isin(target_groups)].copy()
                subset = subset.dropna(subset=["employment_total"])

                if subset.empty or subset["employment_total"].sum() == 0:
                    shares = {target_groups[0]: 0.5, target_groups[1]: 0.5}
                else:
                    shares = (
                        subset.set_index("branch_10a3")["employment_total"] /
                        subset["employment_total"].sum()
                    ).to_dict()

                for g in target_groups:
                    out = row.to_dict()
                    out["branch_10a3"] = g
                    out["allocation_share"] = shares.get(g, 0.0)
                    base_rows.append(out)

        base_10a3 = pd.DataFrame(base_rows)
        for c in ["Production", "NVA"]:
            base_10a3[c] = pd.to_numeric(base_10a3[c], errors="coerce") * base_10a3["allocation_share"]

        base_10a3 = (
            base_10a3.groupby("branch_10a3", as_index=False)[["Production", "NVA"]]
            .sum(min_count=1)
        )

        exposure = exp_10a3.merge(base_10a3, on="branch_10a3", how="left")
        exposure = exposure.merge(dst_emp, on="branch_10a3", how="left")

        exposure["exports_over_production"] = np.where(
            exposure["Production"] > 0,
            exposure["exports_total"] / exposure["Production"],
            np.nan
        )

        exposure["dest_linked_NVA"] = np.where(
            exposure["Production"] > 0,
            exposure["NVA"] * exposure["exports_total"] / exposure["Production"],
            np.nan
        )

        exposure["dest_linked_employment_total"] = np.where(
            exposure["Production"] > 0,
            exposure["employment_total"] * exposure["exports_total"] / exposure["Production"],
            np.nan
        )

        exposure["dest_linked_employment_wage_earners"] = np.where(
            exposure["Production"] > 0,
            exposure["employment_wage_earners"] * exposure["exports_total"] / exposure["Production"],
            np.nan
        )

        exposure = exposure.sort_values(["exp_country_name", "branch_10a3"]).reset_index(drop=True)

    return exp_10a3, exposure

# =========================================================
# EXPORTER TOTALS
# =========================================================

def build_exporter_totals(isic_exports):
    out = (
        isic_exports.groupby(["exp_country_code", "exp_country_name", "dest_country_code", "dest_country_name"], as_index=False)
        [["exports_intermediate", "exports_final", "exports_total"]]
        .sum(min_count=1)
    )
    out = out.sort_values("exp_country_name").reset_index(drop=True)
    return out

# =========================================================
# MAIN
# =========================================================

def main():
    print("Reading ADB MRIO ...")
    Z, F, sector_lookup, ctry_code_to_name, ctry_name_to_code = read_mrio_blocks(
        IN_PATH, SHEET_MRIO, SHEET_LEGEND
    )

    print("Extracting exports to USA from Denmark and euro-area countries ...")
    isic_exports, ea_present_names, ea_present_codes = extract_exports_to_usa_by_exporter(
        Z, F, sector_lookup, ctry_code_to_name, ctry_name_to_code
    )

    print("Loading existing workbook mapping ...")
    mapping_small, nace_small = load_existing_mapping_and_nace(TARGET_FILE)

    print("Building NACE export table ...")
    nace_exports = build_nace_exports(isic_exports, mapping_small)

    print("Loading optional DST employment ...")
    dst_emp = load_dst_10a3_employment(DST_EMPLOYMENT_FILE)

    print("Building 10a3 export tables ...")
    exp_10a3, exposure_10a3 = build_10a3_exports_and_exposure(
        nace_exports=nace_exports,
        nace_small=nace_small,
        dst_emp=dst_emp
    )

    exporter_totals = build_exporter_totals(isic_exports)

    sheet_frames = {
        "EXP_USA_ISIC_by_exporter": isic_exports,
        "EXP_USA_NACE_by_exporter": nace_exports,
        "EXP_USA_10a3_by_exporter": exp_10a3,
        "EXP_USA_exporter_totals": exporter_totals
    }

    if exposure_10a3 is not None:
        sheet_frames["EXP_USA_10a3_exposure"] = exposure_10a3

    replace_sheets_in_existing_workbook(TARGET_FILE, sheet_frames)

    print("Updated existing workbook:", TARGET_FILE)
    print("Added/replaced sheets:")
    for s in sheet_frames.keys():
        print(" -", s)

    print("Euro-area countries present in MRIO:", ea_present_names)
    if "Slovakia" not in ea_present_names:
        print("NOTE: Slovakia not present in this MRIO file; EA aggregate excludes missing members.")

if __name__ == "__main__":
    main()