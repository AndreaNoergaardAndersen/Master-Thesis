import re
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# -----------------------------
# USER INPUT
# -----------------------------
IN_PATH = r"\\srv9dnbfil002\userhome\kons-ana\Desktop\Data\ADB-MRIO-2024-August 2025.xlsx"
SHEET_MRIO = None
SHEET_LEGEND = "Legend"

# Existing workbook to amend
out_folder = os.path.dirname(IN_PATH)
TARGET_FILE = os.path.join(out_folder, "adb_mrio_intermediate_imports_long.xlsx")

# -----------------------------
# Helpers
# -----------------------------
EA20 = {
    "Austria","Belgium","Croatia","Cyprus","Estonia","Finland","France","Germany",
    "Greece","Ireland","Italy","Latvia","Lithuania","Luxembourg","Malta","Netherlands",
    "Portugal","Slovakia","Slovenia","Spain"
}

EA_CODE = "EA"
EA_NAME = "Euro area (EA20)"

SCENARIO_COUNTRY = "EA_sourced_intermediates_by_country_incl_self"
SCENARIO_TOTAL = "EA_sourced_intermediates_total_EA_incl_self"

NEW_SHEETS = [
    "EA_detailed_Z_long",
    "EA_collapsed_importer_ind",
    "EA_country_sector_share",
    "EA_total_sector_share",
    "EA_overall_share"
]

def detect_mrio_sheet(xls: pd.ExcelFile) -> str:
    for s in xls.sheet_names:
        if s.lower() != "legend":
            return s
    raise ValueError("Could not find the MRIO sheet (other than 'Legend').")

def find_header_row(df0: pd.DataFrame) -> tuple[int, int]:
    for i in range(min(80, df0.shape[0])):
        row = df0.iloc[i, :].astype(str).tolist()
        hits = sum(bool(re.fullmatch(r"c\d+", x.strip())) for x in row if x not in ("nan", "None"))
        if hits >= 20:
            for j, x in enumerate(row):
                if x.strip() == "c1":
                    return i, j
    raise ValueError("Could not find sector-code header row (c1,c2,...) in the first 80 rows.")

def is_country_code(x) -> bool:
    if not isinstance(x, str):
        return False
    x = x.strip()
    return bool(re.fullmatch(r"[A-Z]{3}|RoW", x))

def read_intermediate_block(in_path: str, sheet_mrio: str, sheet_legend: str):
    xls = pd.ExcelFile(in_path, engine="openpyxl")
    if sheet_mrio is None:
        sheet_mrio = detect_mrio_sheet(xls)

    legend = pd.read_excel(in_path, sheet_name=sheet_legend, engine="openpyxl")
    legend = legend.rename(columns={legend.columns[0]: "Code", legend.columns[1]: "Country"})
    ctry_code_to_name = dict(zip(legend["Code"].astype(str), legend["Country"].astype(str)))
    ctry_name_to_code = dict(zip(legend["Country"].astype(str), legend["Code"].astype(str)))

    df0 = pd.read_excel(in_path, sheet_name=sheet_mrio, header=None, engine="openpyxl")

    sector_code_row, col_start = find_header_row(df0)
    country_code_row = sector_code_row - 1
    sector_name_row = sector_code_row - 2
    data_start_row = sector_code_row + 1

    col_countries = df0.iloc[country_code_row, col_start:].tolist()
    col_sectors = df0.iloc[sector_code_row, col_start:].tolist()

    int_cols = []
    for k, (cc, sc) in enumerate(zip(col_countries, col_sectors)):
        if not (isinstance(sc, str) and re.fullmatch(r"c\d+", sc.strip())):
            break
        if not is_country_code(str(cc)):
            break
        int_cols.append(col_start + k)

    if len(int_cols) < 200:
        raise ValueError(f"Found only {len(int_cols)} intermediate columns — layout may differ.")

    sector_codes_35 = [str(x).strip() for x in df0.iloc[sector_code_row, col_start:col_start+35].tolist()]
    sector_names_35 = [str(x).strip() for x in df0.iloc[sector_name_row, col_start:col_start+35].tolist()]
    sector_lookup = dict(zip(sector_codes_35, sector_names_35))

    row_sectorname_col = col_start - 3
    row_country_col = col_start - 2
    row_sectorcode_col = col_start - 1

    rows = []
    row_keys = []
    for i in range(data_start_row, df0.shape[0]):
        cc = df0.iat[i, row_country_col]
        sc = df0.iat[i, row_sectorcode_col]
        sn = df0.iat[i, row_sectorname_col]

        if not is_country_code(str(cc)):
            break
        if not (isinstance(sc, str) and re.fullmatch(r"c\d+", sc.strip())):
            break
        if pd.isna(sn):
            break

        rows.append(i)
        row_keys.append((str(cc).strip(), str(sc).strip()))

    if len(rows) < 200:
        raise ValueError(f"Found only {len(rows)} intermediate rows — layout may differ.")

    Z = df0.iloc[rows, int_cols].copy()
    Z = Z.apply(pd.to_numeric, errors="coerce")

    Z.index = pd.MultiIndex.from_tuples(row_keys, names=["exp_country", "exp_sector"])
    col_keys = [(str(df0.iat[country_code_row, j]).strip(), str(df0.iat[sector_code_row, j]).strip()) for j in int_cols]
    Z.columns = pd.MultiIndex.from_tuples(col_keys, names=["imp_country", "imp_sector"])

    return Z, sector_lookup, ctry_code_to_name, ctry_name_to_code

def to_long(Z: pd.DataFrame, sector_lookup: dict, ctry_code_to_name: dict,
            exporters: list[str], importers: list[str], label: str) -> pd.DataFrame:
    sub = Z.loc[
        Z.index.get_level_values("exp_country").isin(exporters),
        Z.columns.get_level_values("imp_country").isin(importers)
    ]

    long = sub.stack(["imp_country", "imp_sector"]).reset_index()
    long = long.rename(columns={0: "value"})
    long["scenario"] = label

    long["exp_country_name"] = long["exp_country"].map(ctry_code_to_name)
    long["imp_country_name"] = long["imp_country"].map(ctry_code_to_name)
    long["exp_sector_name"] = long["exp_sector"].map(sector_lookup)
    long["imp_sector_name"] = long["imp_sector"].map(sector_lookup)
    return long

def collapse_by_importer_industry(long_df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    out = (
        long_df
        .groupby(group_cols, as_index=False)["value"]
        .sum(min_count=1)
    )
    out["view"] = "by_importer_industry"
    return out

def add_area_aggregate_long(long_df: pd.DataFrame,
                            area_code: str,
                            area_name: str,
                            member_exporter_codes: list[str] | None = None,
                            member_importer_codes: list[str] | None = None) -> pd.DataFrame:
    """
    Appends rows where:
      - exp_country is replaced by area_code and summed over member_exporter_codes
      - imp_country is replaced by area_code and summed over member_importer_codes
      - both replaced if both lists provided
    Keeps sector detail.
    """
    df = long_df.copy()
    df["value"] = pd.to_numeric(df["value"], errors="coerce")

    out_parts = [df]

    keys = [
        "exp_country", "exp_sector", "imp_country", "imp_sector", "scenario", "view",
        "exp_country_name", "imp_country_name", "exp_sector_name", "imp_sector_name"
    ]

    if member_exporter_codes is not None:
        exp = df[df["exp_country"].isin(member_exporter_codes)].copy()
        if not exp.empty:
            exp["exp_country"] = area_code
            exp["exp_country_name"] = area_name
            exp = exp.groupby(keys, as_index=False)["value"].sum(min_count=1)
            out_parts.append(exp)

    if member_importer_codes is not None:
        imp = df[df["imp_country"].isin(member_importer_codes)].copy()
        if not imp.empty:
            imp["imp_country"] = area_code
            imp["imp_country_name"] = area_name
            imp = imp.groupby(keys, as_index=False)["value"].sum(min_count=1)
            out_parts.append(imp)

    if (member_exporter_codes is not None) and (member_importer_codes is not None):
        both = df[
            df["exp_country"].isin(member_exporter_codes) &
            df["imp_country"].isin(member_importer_codes)
        ].copy()
        if not both.empty:
            both["exp_country"] = area_code
            both["imp_country"] = area_code
            both["exp_country_name"] = area_name
            both["imp_country_name"] = area_name
            both = both.groupby(keys, as_index=False)["value"].sum(min_count=1)
            out_parts.append(both)

    return pd.concat(out_parts, ignore_index=True)

def replace_sheets_in_existing_workbook(target_file: str, sheet_frames: dict[str, pd.DataFrame]):
    """
    Opens existing workbook, deletes sheets if they already exist, and writes replacements.
    """
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"Target workbook not found: {target_file}")

    wb = load_workbook(target_file)

    for s in sheet_frames.keys():
        if s in wb.sheetnames:
            del wb[s]

    wb.save(target_file)
    wb.close()

    with pd.ExcelWriter(
        target_file,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        for sheet_name, df in sheet_frames.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# -----------------------------
# Main
# -----------------------------
def main():
    Z, sector_lookup, ctry_code_to_name, ctry_name_to_code = read_intermediate_block(
        IN_PATH, SHEET_MRIO, SHEET_LEGEND
    )

    present_countries = set(ctry_name_to_code.keys())
    ea_present_names = sorted(EA20.intersection(present_countries))
    ea_present_codes = [ctry_name_to_code[n] for n in ea_present_names]

    if len(ea_present_codes) == 0:
        raise ValueError("No euro-area countries found in the MRIO file.")

    # =========================================================
    # 1. Detailed EA -> EA intermediate flows, incl. own country
    # =========================================================
    long_ea_to_ea = to_long(
        Z=Z,
        sector_lookup=sector_lookup,
        ctry_code_to_name=ctry_code_to_name,
        exporters=ea_present_codes,
        importers=ea_present_codes,
        label=SCENARIO_COUNTRY
    )
    long_ea_to_ea["view"] = "by_exporter_and_importer_industry"

    # Add EA pseudo-country aggregates:
    # - EA as exporter
    # - EA as importer
    # - EA -> EA
    long_ea_to_ea_with_area = add_area_aggregate_long(
        long_df=long_ea_to_ea,
        area_code=EA_CODE,
        area_name=EA_NAME,
        member_exporter_codes=ea_present_codes,
        member_importer_codes=ea_present_codes
    )

    # =========================================================
    # 2. Collapsed by importer industry
    # =========================================================
    collapsed_ea_to_ea = collapse_by_importer_industry(
        long_ea_to_ea,
        group_cols=[
            "scenario",
            "imp_country",
            "imp_country_name",
            "imp_sector",
            "imp_sector_name"
        ]
    )

    # Add EA aggregate importer row
    collapsed_ea_importer = collapsed_ea_to_ea.copy()
    collapsed_ea_importer = collapsed_ea_importer[
        collapsed_ea_importer["imp_country"].isin(ea_present_codes)
    ].copy()

    collapsed_ea_importer["imp_country"] = EA_CODE
    collapsed_ea_importer["imp_country_name"] = EA_NAME

    collapsed_ea_importer = (
        collapsed_ea_importer
        .groupby(
            ["scenario", "imp_country", "imp_country_name", "imp_sector", "imp_sector_name", "view"],
            as_index=False
        )["value"]
        .sum(min_count=1)
    )

    collapsed_ea_to_ea_with_area = pd.concat(
        [collapsed_ea_to_ea, collapsed_ea_importer],
        ignore_index=True
    )

    # =========================================================
    # 3. Country-sector shares
    # Numerator: EA-sourced intermediates incl self
    # Denominator: total intermediates from all source countries
    # =========================================================

    # Numerator by importer country-sector
    numer_country = (
        long_ea_to_ea
        .groupby(
            ["imp_country", "imp_country_name", "imp_sector", "imp_sector_name"],
            as_index=False
        )["value"]
        .sum(min_count=1)
        .rename(columns={"value": "EA_input_from_EA_incl_self"})
    )

    # Denominator from full Z, all exporters -> EA importers
    all_to_ea = to_long(
        Z=Z,
        sector_lookup=sector_lookup,
        ctry_code_to_name=ctry_code_to_name,
        exporters=sorted(set(Z.index.get_level_values("exp_country"))),
        importers=ea_present_codes,
        label="total_intermediate_inputs"
    )
    all_to_ea["view"] = "by_exporter_and_importer_industry"

    denom_country = (
        all_to_ea
        .groupby(
            ["imp_country", "imp_country_name", "imp_sector", "imp_sector_name"],
            as_index=False
        )["value"]
        .sum(min_count=1)
        .rename(columns={"value": "intermediate_input_total"})
    )

    country_sector_share = numer_country.merge(
        denom_country,
        on=["imp_country", "imp_country_name", "imp_sector", "imp_sector_name"],
        how="outer"
    )

    country_sector_share["scenario"] = SCENARIO_COUNTRY
    country_sector_share["EA_input_share_in_total_intermediates"] = (
        country_sector_share["EA_input_from_EA_incl_self"] /
        country_sector_share["intermediate_input_total"]
    )

    country_sector_share = country_sector_share[
        [
            "scenario",
            "imp_country",
            "imp_country_name",
            "imp_sector",
            "imp_sector_name",
            "EA_input_from_EA_incl_self",
            "intermediate_input_total",
            "EA_input_share_in_total_intermediates"
        ]
    ].rename(columns={
        "imp_country": "Country_code",
        "imp_country_name": "Import_country",
        "imp_sector": "ISIC_sector_code",
        "imp_sector_name": "ISIC_sector"
    })

    # =========================================================
    # 4. EA total by sector
    # =========================================================
    numer_total = (
        numer_country
        .groupby(["imp_sector", "imp_sector_name"], as_index=False)["EA_input_from_EA_incl_self"]
        .sum(min_count=1)
    )

    denom_total = (
        denom_country
        .groupby(["imp_sector", "imp_sector_name"], as_index=False)["intermediate_input_total"]
        .sum(min_count=1)
    )

    total_sector_share = numer_total.merge(
        denom_total,
        on=["imp_sector", "imp_sector_name"],
        how="outer"
    )

    total_sector_share["scenario"] = SCENARIO_TOTAL
    total_sector_share["EA_input_share_in_total_intermediates"] = (
        total_sector_share["EA_input_from_EA_incl_self"] /
        total_sector_share["intermediate_input_total"]
    )

    total_sector_share = total_sector_share[
        [
            "scenario",
            "imp_sector",
            "imp_sector_name",
            "EA_input_from_EA_incl_self",
            "intermediate_input_total",
            "EA_input_share_in_total_intermediates"
        ]
    ].rename(columns={
        "imp_sector": "ISIC_sector_code",
        "imp_sector_name": "ISIC_sector"
    })

    # =========================================================
    # 5. Overall EA share
    # =========================================================
    overall_ea = total_sector_share["EA_input_from_EA_incl_self"].sum(min_count=1)
    overall_total = total_sector_share["intermediate_input_total"].sum(min_count=1)
    overall_share = overall_ea / overall_total if pd.notna(overall_total) and overall_total != 0 else np.nan

    overall_share_df = pd.DataFrame({
        "scenario": [SCENARIO_TOTAL],
        "EA_input_from_EA_incl_self": [overall_ea],
        "intermediate_input_total": [overall_total],
        "EA_input_share_in_total_intermediates": [overall_share]
    })

    # =========================================================
    # 6. Write into existing workbook
    # =========================================================
    sheet_frames = {
        "EA_detailed_Z_long": long_ea_to_ea_with_area,
        "EA_collapsed_importer_ind": collapsed_ea_to_ea_with_area,
        "EA_country_sector_share": country_sector_share,
        "EA_total_sector_share": total_sector_share,
        "EA_overall_share": overall_share_df
    }

    replace_sheets_in_existing_workbook(TARGET_FILE, sheet_frames)

    print("Updated existing workbook:", TARGET_FILE)
    print("Added/replaced sheets:")
    for s in sheet_frames.keys():
        print(" -", s)
    print("EA countries present in file:", ea_present_names)

    if "Slovakia" not in ea_present_names:
        print("NOTE: Slovakia not present in this MRIO file; EA aggregate excludes missing members.")

if __name__ == "__main__":
    main()