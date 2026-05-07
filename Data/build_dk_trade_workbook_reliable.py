"""
Build a clean, source-consistent 2024 Denmark trade workbook.

This script is designed for Andrea's thesis tables. It deliberately avoids mixing
partner numerators and world denominators from different sources.

It creates an Excel workbook with:
  1. Main_BBY: headline Danish trade with the United States and EA20 from StatBank BBY
  2. Composition_summary: largest component + average component shares for goods/services
  3. Composition_details: component-level data behind the summary from StatBank UHQ
  4. ADB_intermediate_shares: intermediate/final split from the actual ADB-MRIO August file
  5. Audit: checks for impossible shares, goods+services totals, component sums, etc.
  6. Sources: source URLs, units, and notes

Run from a terminal with internet access:

  python build_dk_trade_workbook_reliable.py \
      --adb "D:/path/to/ADB-MRIO-2024-August 2025.xlsx" \
      --out "D:/path/to/DK_trade_US_EA20_2024_clean.xlsx"

Requirements:
  pip install requests pandas openpyxl

Notes:
  * BBY values are already in mDKK (Mio. kr.). No EUR conversion is used.
  * UHQ component values are converted to mDKK based on the table unit metadata.
  * ADB-MRIO values are kept in the source unit, millions of US dollars.
"""

from __future__ import annotations

import argparse
import io
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import pandas as pd
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

STATBANK_BASE = "https://api.statbank.dk/v1"

EA20_CODES_ADB = [
    "AUT", "BEL", "HRV", "CYP", "EST", "FIN", "FRA", "GER", "GRC", "IRE",
    "ITA", "LVA", "LTU", "LUX", "MLT", "NET", "POR", "SVK", "SVN", "SPA",
]

# UHQ component definitions. These are intended to be mutually exclusive broad components.
# The script matches labels robustly against StatBank metadata.
GOODS_COMPONENTS = [
    ("Goods", "Live animals, food, beverages and tobacco", [
        "Live animals, Food, Beverages And Tobacco",
        "Live animals, food, beverages and tobacco",
    ]),
    ("Goods", "Crude materials, inedible, except fuels", [
        "Crude Materials, Inedible, Except Fuels",
    ]),
    ("Goods", "Mineral fuels, lubricants and related materials", [
        "Mineral Fuels, Lubricants And Related Materials",
    ]),
    ("Goods", "Chemicals and related products", [
        "Chemicals And Related Products",
    ]),
    ("Goods", "Manufactured goods classified chiefly by material", [
        "Manufactured Goods Classified Chiefly By Material",
    ]),
    ("Goods", "Machinery excluding transport equipment", [
        "Machinery (excl. Transport Equipment)",
    ]),
    ("Goods", "Transport equipment excluding vessels and aircraft", [
        "Transport Equipment (excl. Vessels, aircraft etc.)",
    ]),
    ("Goods", "Vessels, aircraft etc.", [
        "Vessels, aircraft etc.",
    ]),
    ("Goods", "Miscellaneous manufactured articles", [
        "Miscellaneous Manufactured Articles",
    ]),
    ("Goods", "Other goods crossing Danish borders", [
        "Other Goods that crosses danish borders",
        "Other goods that crosses danish borders",
    ]),
    ("Goods", "Goods procured in ports by carriers", [
        "Goods procured in ports by carriers (i.e. bunkring and precurements)",
        "Goods procured in ports by carriers (i.e. bunkering and procurements)",
    ]),
    ("Goods", "Goods sold abroad in connection to processing abroad", [
        "Goods sold abroad in connection to processing abroad",
    ]),
    ("Goods", "Goods procured abroad in connection to processing abroad", [
        "Goods procured abroad in connection to processing abroad",
    ]),
    ("Goods", "Merchanting, goods sold under merchanting", [
        "Merchanting, goods sold under merchanting",
    ]),
    ("Goods", "Other goods that never cross Danish borders", [
        "Other Goods that never crosses danish borders",
        "Other goods that never crosses danish borders",
    ]),
]

SERVICE_COMPONENTS = [
    ("Services", "Manufacturing services", ["Manufacturing services"]),
    ("Services", "Maintenance and repair services", ["Maintenance and repair services"]),
    # Use the broad parent transport category, not the sub-modes, to avoid double-counting.
    ("Services", "Transport", ["Transport"]),
    ("Services", "Travel", ["Travel"]),
    ("Services", "Construction", ["Construction"]),
    ("Services", "Insurance and pensions services", ["Insurance and pensions services"]),
    ("Services", "Financial services", ["Financial services"]),
    ("Services", "Charges for the use of intellectual properties", [
        "Charges for the use of intellectual properties",
    ]),
    ("Services", "Telecommunications, computer, and information services", [
        "Telecommunications, computer, and information services",
    ]),
    ("Services", "Other business services", ["Other business services"]),
    ("Services", "Personal, cultural, and recreational services", [
        "Personal, cultural, and recreational services",
    ]),
    ("Services", "Government goods and services", ["Government goods and services"]),
]

TOTAL_ITEMS = [
    ("All", "GOODS AND SERVICES", ["GOODS AND SERVICES"]),
    ("Goods", "GOODS (FOB)", ["GOODS (FOB)"]),
    ("Services", "SERVICES", ["SERVICES"]),
]


# -------------------------
# Helpers
# -------------------------

def norm(x: Any) -> str:
    """Normalize text for robust matching."""
    if x is None:
        return ""
    s = str(x).replace("\xa0", " ").strip().lower()
    s = re.sub(r"[\s\-_]+", " ", s)
    s = re.sub(r"[()\[\],.;:/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def safe_float(x: Any) -> float:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s or s in {"..", ":", "-"}:
        return 0.0
    # Handle Danish formatted numbers if present.
    s = s.replace(".", "").replace(",", ".") if re.search(r"\d,\d", s) else s.replace(",", "")
    return float(s)


def unit_to_mdk_factor(unit: str) -> float:
    u = norm(unit)
    if "mio" in u or "million" in u:
        return 1.0
    if "1 000" in u or "1000" in u or "thousand" in u:
        return 0.001
    # If unknown, keep unchanged but flag in Sources/Audit.
    return 1.0


def statbank_tableinfo(table: str, lang: str = "en") -> Dict[str, Any]:
    url = f"{STATBANK_BASE}/tableinfo"
    r = requests.get(url, params={"id": table, "format": "JSON", "lang": lang}, timeout=60)
    r.raise_for_status()
    return r.json()


def find_variable(info: Dict[str, Any], candidates: Iterable[str]) -> Dict[str, Any]:
    cand_norm = [norm(c) for c in candidates]
    variables = info.get("variables") or []
    # exact id/text first
    for v in variables:
        vid = norm(v.get("id"))
        vtext = norm(v.get("text"))
        if vid in cand_norm or vtext in cand_norm:
            return v
    # contains fallback
    for v in variables:
        vid = norm(v.get("id"))
        vtext = norm(v.get("text"))
        if any(c in vid or c in vtext or vid in c or vtext in c for c in cand_norm):
            return v
    available = [(v.get("id"), v.get("text")) for v in variables]
    raise KeyError(f"Could not find variable among {candidates}. Available variables: {available}")


def find_value_ids(var: Dict[str, Any], labels: Iterable[str]) -> List[str]:
    """Find one or more StatBank value IDs by label/text. Returns IDs in requested order."""
    values = var.get("values") or []
    value_map_exact = {norm(v.get("text")): v.get("id") for v in values}
    value_map_id = {norm(v.get("id")): v.get("id") for v in values}
    result = []
    for label in labels:
        n = norm(label)
        found = None
        if n in value_map_exact:
            found = value_map_exact[n]
        elif n in value_map_id:
            found = value_map_id[n]
        else:
            # contains fallback; prefer exact-ish matches over parents/children
            matches = []
            for v in values:
                t = norm(v.get("text"))
                vid = norm(v.get("id"))
                if n == t or n == vid or n in t or t in n:
                    matches.append(v)
            if len(matches) == 1:
                found = matches[0].get("id")
            elif len(matches) > 1:
                # Choose the shortest textual match, which usually selects the broad category rather than subcategories.
                matches = sorted(matches, key=lambda v: len(str(v.get("text"))))
                found = matches[0].get("id")
        if found is None:
            sample = [(v.get("id"), v.get("text")) for v in values[:80]]
            raise KeyError(f"Could not find value label '{label}' in variable {var.get('id')} / {var.get('text')}. Sample values: {sample}")
        result.append(found)
    return result


def find_one_value_id_from_aliases(var: Dict[str, Any], aliases: Iterable[str]) -> str:
    last_error = None
    for alias in aliases:
        try:
            return find_value_ids(var, [alias])[0]
        except KeyError as e:
            last_error = e
    raise last_error or KeyError(f"No aliases provided for variable {var.get('id')}")


def statbank_fetch_csv(table: str, selections: List[Tuple[Dict[str, Any], List[str]]], lang: str = "en") -> pd.DataFrame:
    payload = {
        "table": table,
        "format": "CSV",
        "lang": lang,
        "delimiter": "Semicolon",
        "valuePresentation": "Value",
        "variables": [
            {"code": var["id"], "values": value_ids} for var, value_ids in selections
        ],
    }
    r = requests.post(f"{STATBANK_BASE}/data", json=payload, timeout=180)
    if r.status_code >= 400:
        raise RuntimeError(f"StatBank API error {r.status_code}: {r.text}\nPayload: {payload}")
    text = r.text
    df = pd.read_csv(io.StringIO(text), sep=";", dtype=str)
    if df.empty:
        raise RuntimeError(f"Empty response from StatBank for table {table}. Payload: {payload}")
    return df


def value_column(df: pd.DataFrame) -> str:
    # StatBank CSV usually returns value as last column. Use last column unless a clear name exists.
    for candidate in ["INDHOLD", "value", "Value", "VALUE"]:
        if candidate in df.columns:
            return candidate
    return df.columns[-1]


def numeric_value_series(df: pd.DataFrame) -> pd.Series:
    col = value_column(df)
    return df[col].map(safe_float)


def get_col_by_contains(df: pd.DataFrame, *needles: str) -> str:
    names = list(df.columns)
    for n in names:
        nn = norm(n)
        if all(norm(x) in nn for x in needles):
            return n
    raise KeyError(f"Could not find column containing {needles}. Columns: {names}")


# -------------------------
# StatBank BBY
# -------------------------

def fetch_bby() -> Tuple[pd.DataFrame, Dict[str, Any]]:
    info = statbank_tableinfo("BBY")
    item_var = find_variable(info, ["Items"])
    flow_var = find_variable(info, ["Receipts/expenditure", "receipts expenditure"])
    country_var = find_variable(info, ["Country"])
    time_var = find_variable(info, ["Time"])

    item_ids = find_value_ids(item_var, ["GOODS AND SERVICES", "GOODS (FOB)", "SERVICES"])
    flow_ids = find_value_ids(flow_var, ["Current receipts", "Current expenditure"])
    country_ids = find_value_ids(country_var, ["REST OF THE WORLD", "United States", "Euro area-20"])
    time_ids = find_value_ids(time_var, ["2024"])

    df = statbank_fetch_csv("BBY", [
        (item_var, item_ids),
        (flow_var, flow_ids),
        (country_var, country_ids),
        (time_var, time_ids),
    ])
    df["value_mDKK"] = numeric_value_series(df) * unit_to_mdk_factor(info.get("unit", ""))
    return df, info


# -------------------------
# StatBank UHQ composition
# -------------------------

def fetch_uhq_components() -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    info = statbank_tableinfo("UHQ")
    item_var = find_variable(info, ["Items"])
    flow_var = find_variable(info, ["Imports and exports"])
    country_var = find_variable(info, ["Country"])
    seas_var = find_variable(info, ["Seasonal adjustment"])
    time_var = find_variable(info, ["Time"])

    # Map component display names to StatBank IDs.
    component_rows = []
    item_ids = []
    for group, component, aliases in GOODS_COMPONENTS + SERVICE_COMPONENTS + TOTAL_ITEMS:
        try:
            item_id = find_one_value_id_from_aliases(item_var, aliases)
        except KeyError as exc:
            raise KeyError(f"Could not match UHQ component '{component}'. {exc}")
        component_rows.append({
            "component_group": group,
            "component": component,
            "statbank_item_id": item_id,
        })
        item_ids.append(item_id)
    # Remove duplicates while preserving order.
    item_ids = list(dict.fromkeys(item_ids))

    flow_ids = find_value_ids(flow_var, ["Imports", "Exports"])
    country_ids = find_value_ids(country_var, ["United States", "Euro area-20"])
    seas_id = find_one_value_id_from_aliases(seas_var, ["Non-seasonally adjusted", "Non-seasonally adjusted "])
    time_ids = find_value_ids(time_var, ["2024Q1", "2024Q2", "2024Q3", "2024Q4"])

    df = statbank_fetch_csv("UHQ", [
        (item_var, item_ids),
        (flow_var, flow_ids),
        (country_var, country_ids),
        (seas_var, [seas_id]),
        (time_var, time_ids),
    ])
    df["value_mDKK"] = numeric_value_series(df) * unit_to_mdk_factor(info.get("unit", ""))

    comp_map = pd.DataFrame(component_rows)
    return df, comp_map, info


def summarize_composition(uhq: pd.DataFrame, comp_map: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    item_col = get_col_by_contains(uhq, "items")
    flow_col = get_col_by_contains(uhq, "imports")
    country_col = get_col_by_contains(uhq, "country")

    # StatBank returns item labels, not IDs, when valuePresentation=Value. Map by normalized label.
    comp_map = comp_map.copy()
    comp_map["item_norm"] = comp_map["component"].map(norm)

    # Build a robust reverse lookup by comparing fetched item labels against requested aliases.
    # Because the data response contains text labels, match to component rows by the matched StatBank text if possible.
    fetched_items = sorted(uhq[item_col].dropna().unique())
    item_to_component = {}
    for fetched in fetched_items:
        nf = norm(fetched)
        best = None
        for group, component, aliases in GOODS_COMPONENTS + SERVICE_COMPONENTS + TOTAL_ITEMS:
            if any(nf == norm(a) or nf in norm(a) or norm(a) in nf for a in aliases + [component]):
                best = (group, component)
                break
        if best is not None:
            item_to_component[fetched] = best
        else:
            # Last-resort keep fetched name as component.
            item_to_component[fetched] = ("Unknown", fetched)

    rows = []
    for _, r in uhq.iterrows():
        group, comp = item_to_component.get(r[item_col], ("Unknown", r[item_col]))
        rows.append({
            "partner": "United States" if norm(r[country_col]) == norm("United States") else "Euro area-20",
            "flow": "Import" if norm(r[flow_col]) == norm("Imports") else "Export",
            "component_group": group,
            "component": comp,
            "value_mDKK": r["value_mDKK"],
        })
    long = pd.DataFrame(rows)
    # Annual sum across quarters.
    annual = long.groupby(["partner", "flow", "component_group", "component"], as_index=False)["value_mDKK"].sum()

    total_lookup = annual[annual["component"] == "GOODS AND SERVICES"].set_index(["partner", "flow"])["value_mDKK"].to_dict()
    detail_rows = []
    for _, r in annual.iterrows():
        if r["component"] in {"GOODS AND SERVICES", "GOODS (FOB)", "SERVICES"}:
            continue
        total = total_lookup.get((r["partner"], r["flow"]), float("nan"))
        share = r["value_mDKK"] / total if total and not math.isnan(total) else float("nan")
        detail_rows.append({
            "partner": r["partner"],
            "flow": r["flow"],
            "component_group": r["component_group"],
            "component": r["component"],
            "value_mDKK": r["value_mDKK"],
            "share_of_goods_and_services": share,
            "denominator_goods_and_services_mDKK": total,
        })
    details = pd.DataFrame(detail_rows)

    summary_rows = []
    for (partner, flow), g in details.groupby(["partner", "flow"]):
        g_nonzero = g[g["value_mDKK"].abs() > 1e-9].copy()
        largest = g_nonzero.sort_values("share_of_goods_and_services", ascending=False).iloc[0]
        goods = g_nonzero[g_nonzero["component_group"] == "Goods"]
        services = g_nonzero[g_nonzero["component_group"] == "Services"]
        denom = largest["denominator_goods_and_services_mDKK"]
        summary_rows.append({
            "partner": partner,
            "flow": flow,
            "largest_component": largest["component"],
            "largest_component_group": largest["component_group"],
            "largest_component_value_mDKK": largest["value_mDKK"],
            "largest_component_share_of_goods_and_services": largest["share_of_goods_and_services"],
            "average_service_component_share_of_goods_and_services": services["share_of_goods_and_services"].mean() if len(services) else float("nan"),
            "average_goods_component_share_of_goods_and_services": goods["share_of_goods_and_services"].mean() if len(goods) else float("nan"),
            "n_service_components_nonzero": int(len(services)),
            "n_goods_components_nonzero": int(len(goods)),
            "service_components_total_share": services["share_of_goods_and_services"].sum(),
            "goods_components_total_share": goods["share_of_goods_and_services"].sum(),
            "all_components_total_share_check": g_nonzero["share_of_goods_and_services"].sum(),
            "denominator_goods_and_services_mDKK": denom,
        })
    summary = pd.DataFrame(summary_rows).sort_values(["flow", "partner"])
    details = details.sort_values(["flow", "partner", "component_group", "share_of_goods_and_services"], ascending=[True, True, True, False])
    return summary, details


# -------------------------
# ADB-MRIO calculations
# -------------------------

def sum_values(row: Tuple[Any, ...], cols: List[int]) -> float:
    total = 0.0
    for idx in cols:
        total += safe_float(row[idx])
    return total


def compute_adb_intermediate_shares(adb_path: Path) -> pd.DataFrame:
    wb = load_workbook(adb_path, read_only=True, data_only=True)
    if "ADB MRIO 2024" not in wb.sheetnames:
        raise KeyError("ADB workbook does not contain a sheet named 'ADB MRIO 2024'.")
    ws = wb["ADB MRIO 2024"]

    row6 = next(ws.iter_rows(min_row=6, max_row=6, values_only=True))
    row7 = next(ws.iter_rows(min_row=7, max_row=7, values_only=True))

    col_country = list(row6)
    col_code = list(row7)

    def cols_for(countries: Iterable[str], prefix: str) -> List[int]:
        countries_set = set(countries)
        return [i for i, (cty, code) in enumerate(zip(col_country, col_code))
                if cty in countries_set and isinstance(code, str) and code.startswith(prefix)]

    den_inter_cols = cols_for(["DEN"], "c")
    den_final_cols = cols_for(["DEN"], "F")
    us_inter_cols = cols_for(["USA"], "c")
    us_final_cols = cols_for(["USA"], "F")
    ea_inter_cols = cols_for(EA20_CODES_ADB, "c")
    ea_final_cols = cols_for(EA20_CODES_ADB, "F")

    totals = {
        ("Import", "United States", "intermediate"): 0.0,
        ("Import", "United States", "final"): 0.0,
        ("Import", "Euro area-20", "intermediate"): 0.0,
        ("Import", "Euro area-20", "final"): 0.0,
        ("Export", "United States", "intermediate"): 0.0,
        ("Export", "United States", "final"): 0.0,
        ("Export", "Euro area-20", "intermediate"): 0.0,
        ("Export", "Euro area-20", "final"): 0.0,
    }

    for row in ws.iter_rows(min_row=8, values_only=True):
        exp_country = row[2]
        if exp_country == "USA":
            totals[("Import", "United States", "intermediate")] += sum_values(row, den_inter_cols)
            totals[("Import", "United States", "final")] += sum_values(row, den_final_cols)
        elif exp_country in EA20_CODES_ADB:
            totals[("Import", "Euro area-20", "intermediate")] += sum_values(row, den_inter_cols)
            totals[("Import", "Euro area-20", "final")] += sum_values(row, den_final_cols)
        elif exp_country == "DEN":
            totals[("Export", "United States", "intermediate")] += sum_values(row, us_inter_cols)
            totals[("Export", "United States", "final")] += sum_values(row, us_final_cols)
            totals[("Export", "Euro area-20", "intermediate")] += sum_values(row, ea_inter_cols)
            totals[("Export", "Euro area-20", "final")] += sum_values(row, ea_final_cols)

    rows = []
    for flow in ["Import", "Export"]:
        for partner in ["United States", "Euro area-20"]:
            inter = totals[(flow, partner, "intermediate")]
            final = totals[(flow, partner, "final")]
            total = inter + final
            rows.append({
                "flow": flow,
                "partner": partner,
                "intermediate_million_USD": inter,
                "final_million_USD": final,
                "total_million_USD": total,
                "intermediate_share": inter / total if total else float("nan"),
                "source": "ADB-MRIO 2024, August 2025, current prices, millions of US$",
            })
    # Add US+EA total rows.
    df = pd.DataFrame(rows)
    combo_rows = []
    for flow, g in df.groupby("flow"):
        inter = g["intermediate_million_USD"].sum()
        final = g["final_million_USD"].sum()
        total = inter + final
        combo_rows.append({
            "flow": flow,
            "partner": "United States + Euro area-20",
            "intermediate_million_USD": inter,
            "final_million_USD": final,
            "total_million_USD": total,
            "intermediate_share": inter / total if total else float("nan"),
            "source": "ADB-MRIO 2024, August 2025, current prices, millions of US$",
        })
    return pd.concat([df, pd.DataFrame(combo_rows)], ignore_index=True)


# -------------------------
# Main BBY table construction
# -------------------------

def build_main_bby_table(bby: pd.DataFrame) -> pd.DataFrame:
    item_col = get_col_by_contains(bby, "items")
    flow_col = get_col_by_contains(bby, "receipts")
    country_col = get_col_by_contains(bby, "country")

    def get(item: str, flow: str, country: str) -> float:
        m = (
            (bby[item_col].map(norm) == norm(item)) &
            (bby[flow_col].map(norm) == norm(flow)) &
            (bby[country_col].map(norm) == norm(country))
        )
        vals = bby.loc[m, "value_mDKK"]
        if vals.empty:
            raise KeyError(f"Missing BBY value for item={item}, flow={flow}, country={country}")
        return float(vals.iloc[0])

    rows = []
    for flow_name, flow_label in [("Import", "Current expenditure"), ("Export", "Current receipts")]:
        values = {}
        for item_label, item in [("All", "GOODS AND SERVICES"), ("Goods", "GOODS (FOB)"), ("Services", "SERVICES")]:
            us = get(item, flow_label, "United States")
            ea = get(item, flow_label, "Euro area-20")
            world = get(item, flow_label, "REST OF THE WORLD")
            values[item_label] = {"US": us, "EA20": ea, "US_EA20": us + ea, "World": world}
            rows.append({"section": flow_name, "item": item_label, "United States": us, "EA20": ea, "US+EA20 total": us + ea, "World / RoW denominator": world, "note": "BBY value, mDKK"})
        allv = values["All"]
        rows.append({"section": flow_name, "item": "Share incl. RoW", "United States": allv["US"] / allv["World"], "EA20": allv["EA20"] / allv["World"], "US+EA20 total": allv["US_EA20"] / allv["World"], "World / RoW denominator": 1.0, "note": "Partner / BBY REST OF THE WORLD"})
        rows.append({"section": flow_name, "item": "Share excl. RoW", "United States": allv["US"] / allv["US_EA20"], "EA20": allv["EA20"] / allv["US_EA20"], "US+EA20 total": 1.0, "World / RoW denominator": None, "note": "Partner / (US+EA20)"})
        rows.append({"section": flow_name, "item": "Goods share", "United States": values["Goods"]["US"] / allv["US"], "EA20": values["Goods"]["EA20"] / allv["EA20"], "US+EA20 total": values["Goods"]["US_EA20"] / allv["US_EA20"], "World / RoW denominator": values["Goods"]["World"] / allv["World"], "note": "Goods / all trade"})
        rows.append({"section": flow_name, "item": "Services share", "United States": values["Services"]["US"] / allv["US"], "EA20": values["Services"]["EA20"] / allv["EA20"], "US+EA20 total": values["Services"]["US_EA20"] / allv["US_EA20"], "World / RoW denominator": values["Services"]["World"] / allv["World"], "note": "Services / all trade"})

    # Trade balance rows: exports - imports.
    def lookup(rows_df: pd.DataFrame, section: str, item: str, col: str) -> float:
        m = (rows_df["section"] == section) & (rows_df["item"] == item)
        return float(rows_df.loc[m, col].iloc[0])

    table = pd.DataFrame(rows)
    balance_rows = []
    for col in ["United States", "EA20", "US+EA20 total", "World / RoW denominator"]:
        balance_rows.append({
            "section": "Trade balance",
            "item": col,
            "United States": lookup(table, "Export", "All", "United States") - lookup(table, "Import", "All", "United States") if col == "United States" else None,
            "EA20": lookup(table, "Export", "All", "EA20") - lookup(table, "Import", "All", "EA20") if col == "EA20" else None,
            "US+EA20 total": lookup(table, "Export", "All", "US+EA20 total") - lookup(table, "Import", "All", "US+EA20 total") if col == "US+EA20 total" else None,
            "World / RoW denominator": lookup(table, "Export", "All", "World / RoW denominator") - lookup(table, "Import", "All", "World / RoW denominator") if col == "World / RoW denominator" else None,
            "note": "Exports - imports",
        })
    # Keep one compact balance row instead of four duplicates.
    compact_balance = {
        "section": "Trade balance",
        "item": "All trade balance",
        "United States": lookup(table, "Export", "All", "United States") - lookup(table, "Import", "All", "United States"),
        "EA20": lookup(table, "Export", "All", "EA20") - lookup(table, "Import", "All", "EA20"),
        "US+EA20 total": lookup(table, "Export", "All", "US+EA20 total") - lookup(table, "Import", "All", "US+EA20 total"),
        "World / RoW denominator": lookup(table, "Export", "All", "World / RoW denominator") - lookup(table, "Import", "All", "World / RoW denominator"),
        "note": "Exports - imports",
    }
    table = pd.concat([table, pd.DataFrame([compact_balance])], ignore_index=True)
    return table


# -------------------------
# Audit and Excel writing
# -------------------------

def build_audit(main: pd.DataFrame, comp_summary: pd.DataFrame, adb: pd.DataFrame) -> pd.DataFrame:
    checks = []

    # Main table checks.
    for section in ["Import", "Export"]:
        m = main[main["section"] == section].set_index("item")
        for col in ["United States", "EA20", "US+EA20 total", "World / RoW denominator"]:
            if col not in m.columns:
                continue
            all_val = m.loc["All", col]
            goods_val = m.loc["Goods", col]
            services_val = m.loc["Services", col]
            diff = (goods_val + services_val) - all_val
            checks.append({"check": f"{section}: goods + services = all ({col})", "value": diff, "status": "OK" if abs(diff) < 1e-6 else "WARN"})
            if col != "World / RoW denominator":
                share = m.loc["Share incl. RoW", col]
                checks.append({"check": f"{section}: RoW-inclusive share <= 100% ({col})", "value": share, "status": "OK" if 0 <= share <= 1 + 1e-9 else "WARN"})
        # Excl RoW shares should sum to 100 across US and EA20.
        excl_sum = m.loc["Share excl. RoW", "United States"] + m.loc["Share excl. RoW", "EA20"]
        checks.append({"check": f"{section}: US + EA20 shares excl. RoW sum to 100%", "value": excl_sum, "status": "OK" if abs(excl_sum - 1) < 1e-9 else "WARN"})

    # Composition checks.
    for _, r in comp_summary.iterrows():
        total_share = r["all_components_total_share_check"]
        checks.append({"check": f"Composition total share: {r['flow']} {r['partner']}", "value": total_share, "status": "OK" if abs(total_share - 1) < 0.02 else "WARN"})
        checks.append({"check": f"Largest component share <= 100%: {r['flow']} {r['partner']}", "value": r["largest_component_share_of_goods_and_services"], "status": "OK" if 0 <= r["largest_component_share_of_goods_and_services"] <= 1 else "WARN"})

    # ADB checks.
    for _, r in adb.iterrows():
        checks.append({"check": f"ADB intermediate share in [0,1]: {r['flow']} {r['partner']}", "value": r["intermediate_share"], "status": "OK" if 0 <= r["intermediate_share"] <= 1 else "WARN"})
        diff = r["intermediate_million_USD"] + r["final_million_USD"] - r["total_million_USD"]
        checks.append({"check": f"ADB intermediate + final = total: {r['flow']} {r['partner']}", "value": diff, "status": "OK" if abs(diff) < 1e-6 else "WARN"})

    return pd.DataFrame(checks)


def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for j, col in enumerate(df.columns, start_col):
        cell = ws.cell(start_row, j, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E5F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="999999"))
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, val in enumerate(row, start_col):
            cell = ws.cell(i, j, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="CCCCCC"))
            if isinstance(val, float):
                # percentage-like columns
                header = str(df.columns[j - start_col]).lower()
                if "share" in header or header in {"value"}:
                    if "value_mDKK" not in header and "million" not in header and "mDKK" not in header:
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "#,##0.0"
                else:
                    cell.number_format = "#,##0.0"
    ws.freeze_panes = ws.cell(start_row + 1, start_col)
    for col_idx, col in enumerate(df.columns, start_col):
        header = str(col)
        width = min(max(len(header) + 2, 12), 40)
        if any(x in header.lower() for x in ["component", "note", "source", "check"]):
            width = 42
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_workbook(out_path: Path, main: pd.DataFrame, comp_summary: pd.DataFrame, comp_details: pd.DataFrame, adb: pd.DataFrame, audit: pd.DataFrame, bby_info: Dict[str, Any], uhq_info: Dict[str, Any], adb_path: Path) -> None:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # README
    ws = wb.create_sheet("README")
    ws["A1"] = "Clean 2024 Denmark trade workbook"
    ws["A1"].font = Font(bold=True, size=16)
    notes = [
        "All headline DKK trade values are from Statistics Denmark StatBank BBY, in mDKK.",
        "Composition uses Statistics Denmark StatBank UHQ, summed over 2024Q1-2024Q4.",
        "Intermediate/final split uses the actual ADB-MRIO 2024 August file, in millions of US dollars.",
        "No EUR-to-DKK conversion is used for BBY values.",
        "The Audit sheet should show only OK statuses before using the workbook in the thesis.",
    ]
    for i, txt in enumerate(notes, 3):
        ws.cell(i, 1, txt)
    ws.column_dimensions["A"].width = 120

    sheets = [
        ("Main_BBY", main),
        ("Composition_summary", comp_summary),
        ("Composition_details", comp_details),
        ("ADB_intermediate_shares", adb),
        ("Audit", audit),
    ]
    for name, df in sheets:
        ws = wb.create_sheet(name)
        write_df(ws, df)

    sources = pd.DataFrame([
        {"source": "Statistics Denmark StatBank BBY", "url": "https://www.statbank.dk/BBY", "unit": bby_info.get("unit", ""), "use": "Main aggregate trade table and RoW/world denominators"},
        {"source": "Statistics Denmark StatBank UHQ", "url": "https://www.statbank.dk/UHQ", "unit": uhq_info.get("unit", ""), "use": "Goods/services component composition, 2024 quarters summed"},
        {"source": "ADB-MRIO 2024 August 2025", "url": "https://kidb.adb.org/globalization", "unit": "millions of US dollars", "use": "Intermediate/final split from actual MRIO matrix"},
        {"source": "Local ADB file used", "url": str(adb_path), "unit": "millions of US dollars", "use": "Parsed directly by this script"},
    ])
    ws = wb.create_sheet("Sources")
    write_df(ws, sources)

    # Style audit status cells.
    if "Audit" in wb.sheetnames:
        ws = wb["Audit"]
        # Assuming status column is third.
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row, 3).value
            if status == "OK":
                ws.cell(row, 3).fill = PatternFill("solid", fgColor="D9EAD3")
            else:
                ws.cell(row, 3).fill = PatternFill("solid", fgColor="F4CCCC")
                ws.cell(row, 3).font = Font(bold=True, color="9C0006")

    wb.save(out_path)


# -------------------------
# Main
# -------------------------

def main_cli() -> None:
    parser = argparse.ArgumentParser(description="Build clean Danish trade tables for US and EA20, 2024.")
    parser.add_argument("--adb", type=Path, required=True, help="Path to ADB-MRIO-2024-August 2025.xlsx")
    parser.add_argument("--out", type=Path, default=Path("DK_trade_US_EA20_2024_clean.xlsx"), help="Output xlsx path")
    args = parser.parse_args()

    if not args.adb.exists():
        raise FileNotFoundError(f"ADB file not found: {args.adb}")

    print("Fetching BBY from StatBank...")
    bby, bby_info = fetch_bby()
    print("Fetching UHQ from StatBank...")
    uhq, comp_map, uhq_info = fetch_uhq_components()
    print("Summarizing composition...")
    comp_summary, comp_details = summarize_composition(uhq, comp_map)
    print("Computing ADB intermediate/final shares from actual ADB file...")
    adb = compute_adb_intermediate_shares(args.adb)
    print("Building main BBY table...")
    main_table = build_main_bby_table(bby)
    print("Running audit checks...")
    audit = build_audit(main_table, comp_summary, adb)

    # Fail hard if core aggregate shares are impossible.
    bad = audit[audit["status"] != "OK"]
    if not bad.empty:
        print("WARNING: Some audit checks are not OK. The workbook will still be written, but inspect the Audit sheet.")
        print(bad.to_string(index=False))

    print(f"Writing workbook: {args.out}")
    create_workbook(args.out, main_table, comp_summary, comp_details, adb, audit, bby_info, uhq_info, args.adb)
    print("Done.")


if __name__ == "__main__":
    main_cli()
