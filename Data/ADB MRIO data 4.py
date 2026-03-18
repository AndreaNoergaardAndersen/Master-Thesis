import os
import re
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ============================================================
# USER INPUT
# ============================================================

MRIO_FILE = r"\\srv9dnbfil002\userhome\kons-ana\Desktop\Data\adb_mrio_intermediate_imports_long.xlsx"
EUROSTAT_FILE = r"\\srv9dnbfil002\userhome\kons-ana\Desktop\Data\nama_10_a64__custom_20590952_spreadsheet.xlsx"

YEAR = 2024
DATA_SHEET = "Data"

RAW_SHEET_NAME = "EA raw data"
VEGT_SHEET_NAME = "EA Vægt"

# ============================================================
# EURO AREA COUNTRIES
# ============================================================

EA_COUNTRIES = [
    "Austria", "Belgium", "Croatia", "Cyprus", "Estonia", "Finland", "France", "Germany",
    "Greece", "Ireland", "Italy", "Latvia", "Lithuania", "Luxembourg", "Malta",
    "Netherlands", "Portugal", "Slovakia", "Slovenia", "Spain"
]

EA_REPORTED_LABELS = [
    "Euro area – 20 countries (2023-2025)",
    "Euro area – 20 countries",
    "Euro area"
]

# ============================================================
# TARGET 10a3-STYLE GROUPS (matching your Vægt table style)
# ============================================================

TARGET_GROUPS = {
    "A": "A Landbrug, skovbrug og fiskeri",
    "B": "B Råstofindvinding",
    "C": "C Fremstillingsvirksomhed (industri)",
    "D_E": "D_E Forsyningsvirksomhed",
    "F": "F Bygge- og anlægsvirksomhed",
    "G_I": "G_I Handel og transport mv.",
    "J": "J Information og kommunikation",
    "K": "K Finansiering og forsikring",
    "L": "L Fast ejendom",
    "M_N": "M_N Erhvervsservice",
    "O_Q": "O_Q Offentlig administration, undervisning og sundhed",
    "R_S": "R_S Kultur, fritid og anden service",
}

# ============================================================
# EUROSTAT -> TARGET GROUP MAPPING
# Note:
# - D and E are summed into D_E
# - For G_I we prefer the direct Eurostat aggregate if it exists
#   and only fall back to summing G + H + I if needed
# ============================================================

SECTOR_MAP = {
    "Agriculture, forestry and fishing": ("A", TARGET_GROUPS["A"], "direct"),
    "Mining and quarrying": ("B", TARGET_GROUPS["B"], "direct"),
    "Manufacturing": ("C", TARGET_GROUPS["C"], "direct"),
    "Electricity, gas, steam and air conditioning supply": ("D_E", TARGET_GROUPS["D_E"], "component_D"),
    "Water supply; sewerage, waste management and remediation activities": ("D_E", TARGET_GROUPS["D_E"], "component_E"),
    "Construction": ("F", TARGET_GROUPS["F"], "direct"),
    "Wholesale and retail trade; repair of motor vehicles and motorcycles": ("G_I", TARGET_GROUPS["G_I"], "component_G"),
    "Transportation and storage": ("G_I", TARGET_GROUPS["G_I"], "component_H"),
    "Accommodation and food service activities": ("G_I", TARGET_GROUPS["G_I"], "component_I"),
    "Wholesale and retail trade, transport, accommodation and food service activities": ("G_I", TARGET_GROUPS["G_I"], "direct_aggregate"),
    "Information and communication": ("J", TARGET_GROUPS["J"], "direct"),
    "Financial and insurance activities": ("K", TARGET_GROUPS["K"], "direct"),
    "Real estate activities": ("L", TARGET_GROUPS["L"], "direct"),
    "Professional, scientific and technical activities; administrative and support service activities": ("M_N", TARGET_GROUPS["M_N"], "direct"),
    "Public administration, defence, education, human health and social work activities": ("O_Q", TARGET_GROUPS["O_Q"], "direct"),
    "Arts, entertainment and recreation; other service activities; activities of household and extra-territorial organizations and bodies": ("R_S", TARGET_GROUPS["R_S"], "direct"),
}

INDICATOR_MAP = {
    "Output": "Production EURm",
    "Wages and salaries": "Labour costs EURm"
}

GROUP_ORDER = {
    "A": 1,
    "B": 2,
    "C": 3,
    "D_E": 4,
    "F": 5,
    "G_I": 6,
    "J": 7,
    "K": 8,
    "L": 9,
    "M_N": 10,
    "O_Q": 11,
    "R_S": 12,
}

# ============================================================
# Helpers
# ============================================================

def normalize_text(x):
    if pd.isna(x):
        return None
    return str(x).strip()

def is_blank_row(row):
    return all(pd.isna(x) or str(x).strip() == "" for x in row)

def replace_sheets_in_existing_workbook(target_file: str, sheet_frames: dict[str, pd.DataFrame]):
    if not os.path.exists(target_file):
        raise FileNotFoundError(f"Target workbook not found: {target_file}")

    wb = load_workbook(target_file)

    for sheet_name in sheet_frames.keys():
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    wb.save(target_file)
    wb.close()

    with pd.ExcelWriter(
        target_file,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        for sheet_name, df in sheet_frames.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# ============================================================
# Parse repeated blocks in Eurostat Data sheet
# ============================================================

def parse_eurostat_blocks(xlsx_path: str, data_sheet: str, year: int) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name=data_sheet, header=None, engine="openpyxl")
    rows = df.values.tolist()

    out = []
    i = 0

    while i < len(rows):
        r0 = normalize_text(rows[i][0]) if len(rows[i]) > 0 else None

        if r0 == "Time frequency":
            # Fixed metadata positions in this export format
            unit = normalize_text(rows[i + 1][2]) if i + 1 < len(rows) else None
            nace = normalize_text(rows[i + 2][2]) if i + 2 < len(rows) else None
            indicator = normalize_text(rows[i + 3][2]) if i + 3 < len(rows) else None

            # TIME row
            time_row_idx = i + 5
            if time_row_idx >= len(rows):
                i += 1
                continue

            time_row = rows[time_row_idx]
            year_col = None
            for c, val in enumerate(time_row):
                if normalize_text(val) == str(year):
                    year_col = c
                    break

            if year_col is None:
                i += 1
                continue

            # GEO rows start after "GEO (Labels)"
            k = time_row_idx + 2
            while k < len(rows):
                rr = rows[k]
                rr0 = normalize_text(rr[0]) if len(rr) > 0 else None

                # next block
                if rr0 == "Time frequency":
                    break

                # end of current block
                if is_blank_row(rr):
                    break

                geo = normalize_text(rr[0])
                value = rr[year_col] if year_col < len(rr) else None

                out.append({
                    "year": year,
                    "unit": unit,
                    "nace_eurostat": nace,
                    "indicator_eurostat": indicator,
                    "geo": geo,
                    "value": pd.to_numeric(value, errors="coerce")
                })
                k += 1

            i = k
        else:
            i += 1

    out_df = pd.DataFrame(out)

    if out_df.empty:
        raise ValueError("No Eurostat blocks were parsed from the Data sheet.")

    return out_df

# ============================================================
# Build consolidated raw table
# ============================================================

def build_raw_table(parsed_df: pd.DataFrame) -> pd.DataFrame:
    df = parsed_df.copy()

    # Keep only indicators and sectors used here
    df = df[df["indicator_eurostat"].isin(INDICATOR_MAP.keys())].copy()
    df = df[df["nace_eurostat"].isin(SECTOR_MAP.keys())].copy()

    # Keep EA member states + reported EA aggregate if present
    keep_geo = set(EA_COUNTRIES) | set(EA_REPORTED_LABELS)
    df = df[df["geo"].isin(keep_geo)].copy()

    # Map sectors and indicators
    df["NACE_sector_code"] = df["nace_eurostat"].map(lambda x: SECTOR_MAP[x][0])
    df["NACE_sector_name"] = df["nace_eurostat"].map(lambda x: SECTOR_MAP[x][1])
    df["aggregation_method"] = df["nace_eurostat"].map(lambda x: SECTOR_MAP[x][2])
    df["indicator"] = df["indicator_eurostat"].map(INDICATOR_MAP)

    # Member-state rows
    members = df[df["geo"].isin(EA_COUNTRIES)].copy()
    members["source"] = "member_state"

    # Reported EA aggregate rows, if present
    reported = df[df["geo"].isin(EA_REPORTED_LABELS)].copy()
    if not reported.empty:
        reported = reported.copy()
        reported["geo"] = "EA20 reported aggregate"
        reported["source"] = "reported_EA"

    # Constructed EA sum from member states:
    # - For G_I: prefer direct aggregate if it exists, otherwise use G+H+I components
    # - For D_E: sum D and E components
    def build_ea_sum(group):
        # group is one indicator x target sector
        methods = set(group["aggregation_method"].dropna().tolist())

        # prefer direct aggregate for G_I if present
        if "direct_aggregate" in methods:
            use = group[group["aggregation_method"] == "direct_aggregate"].copy()
        else:
            use = group.copy()

        return use["value"].sum(min_count=1)

    ea_sum = (
        members.groupby(
            ["year", "indicator", "NACE_sector_code", "NACE_sector_name", "geo", "aggregation_method"],
            as_index=False
        )["value"].sum(min_count=1)
    )

    # Now collapse across countries for each sector-indicator using method preference
    ea_sum = (
        ea_sum.groupby(
            ["year", "indicator", "NACE_sector_code", "NACE_sector_name"],
            group_keys=False
        )
        .apply(lambda g: pd.Series({"value": build_ea_sum(g)}))
        .reset_index()
    )

    ea_sum["geo"] = "EA20 sum from member states"
    ea_sum["source"] = "sum_member_states"
    ea_sum["nace_eurostat"] = np.nan
    ea_sum["indicator_eurostat"] = np.nan
    ea_sum["aggregation_method"] = "preferred_sum_rule"

    raw = pd.concat([members, reported, ea_sum], ignore_index=True, sort=False)

    raw = raw[
        [
            "year",
            "geo",
            "source",
            "indicator",
            "nace_eurostat",
            "NACE_sector_code",
            "NACE_sector_name",
            "aggregation_method",
            "value"
        ]
    ].sort_values(
        ["indicator", "NACE_sector_code", "source", "geo"],
        kind="stable"
    )

    return raw

# ============================================================
# Build EA Vægt-like summary
# ============================================================

def build_ea_vaegt(raw_df: pd.DataFrame) -> pd.DataFrame:
    ea = raw_df[raw_df["geo"] == "EA20 sum from member states"].copy()

    prod = (
        ea[ea["indicator"] == "Production EURm"]
        [["NACE_sector_code", "NACE_sector_name", "value"]]
        .rename(columns={"value": "Production EURm"})
    )

    lab = (
        ea[ea["indicator"] == "Labour costs EURm"]
        [["NACE_sector_code", "NACE_sector_name", "value"]]
        .rename(columns={"value": "Labour costs EURm"})
    )

    out = prod.merge(
        lab,
        on=["NACE_sector_code", "NACE_sector_name"],
        how="outer"
    )

    out["Production total EURm"] = out["Production EURm"].sum(min_count=1)
    out["Labour cost total EURm"] = out["Labour costs EURm"].sum(min_count=1)
    out["Production weight"] = out["Production EURm"] / out["Production total EURm"]
    out["Labour cost weight"] = out["Labour costs EURm"] / out["Labour cost total EURm"]

    out["sort_order"] = out["NACE_sector_code"].map(GROUP_ORDER)
    out = out.sort_values(["sort_order", "NACE_sector_code"], kind="stable").drop(columns="sort_order")

    return out

# ============================================================
# Main
# ============================================================

def main():
    parsed = parse_eurostat_blocks(EUROSTAT_FILE, DATA_SHEET, YEAR)
    raw = build_raw_table(parsed)
    ea_vaegt = build_ea_vaegt(raw)

    replace_sheets_in_existing_workbook(
        MRIO_FILE,
        {
            RAW_SHEET_NAME: raw,
            VEGT_SHEET_NAME: ea_vaegt
        }
    )

    print("Updated workbook:", MRIO_FILE)
    print("Added/replaced sheets:")
    print(" -", RAW_SHEET_NAME)
    print(" -", VEGT_SHEET_NAME)

if __name__ == "__main__":
    main()